"""
Table extraction engine v11 - PaddleOCR-first production pipeline.

Pipeline strategy:
1. Turkish-aware preprocessing: perspective correction, deskew, shadow removal,
   illumination normalization, adaptive upscale, CLAHE, and sharpening.
2. Pipeline A: OpenCV grid detection + batched PaddleOCR per cell.
3. Pipeline B: PP-StructureV3 table recognition when grid confidence is low.
4. Pipeline C: Borderless-table reconstruction from PaddleOCR text boxes.
5. Turkish-aware postprocessing, confidence scoring, merged-cell reconstruction,
   and styled Excel export.
"""
from __future__ import annotations

# ── Monkeypatch importlib.metadata to avoid PackageNotFoundError under PyInstaller ──
import importlib.metadata
try:
    _orig_version = importlib.metadata.version
except AttributeError:
    _orig_version = lambda x: "4.10.0"

def _patched_version(distribution_name):
    try:
        return _orig_version(distribution_name)
    except importlib.metadata.PackageNotFoundError:
        name_lower = distribution_name.lower()
        if "opencv" in name_lower:
            import cv2
            return getattr(cv2, "__version__", "4.10.0")
        if "paddle" in name_lower:
            return "3.0.0"
        dep_map = {
            "pyyaml": "yaml",
            "pillow": "PIL",
        }
        import_name = dep_map.get(name_lower, name_lower).replace("-", "_")
        try:
            mod = __import__(import_name)
            return getattr(mod, "__version__", "1.0.0")
        except ImportError:
            raise importlib.metadata.PackageNotFoundError(distribution_name)

importlib.metadata.version = _patched_version

import math
import os
import re
import sys
import unicodedata
from dataclasses import dataclass, field
from difflib import get_close_matches
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import cv2
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage

try:
    import pytesseract
except Exception:  # pragma: no cover - emergency compatibility only
    pytesseract = None


APP_CACHE_DIR = Path(__file__).parent.parent / "app_cache"
APP_CACHE_DIR.mkdir(exist_ok=True)
APP_TESSDATA_DIR = APP_CACHE_DIR / "tessdata"
# ── Redirect Paddle cache BEFORE any paddle import ──
# paddle.dataset.common reads os.path.expanduser('~')/.cache/paddle/dataset
# at import-time and calls os.makedirs().  On some Windows machines ~/.cache
# is a *file* (not a directory), causing a crash.  We detect this and redirect
# HOME / USERPROFILE to a local directory so paddle can create its cache.
_user_home = Path.home()
_home_cache = _user_home / ".cache"
if _home_cache.exists() and not _home_cache.is_dir():
    _paddle_home = APP_CACHE_DIR / "paddle_home"
    _paddle_home.mkdir(parents=True, exist_ok=True)
    (_paddle_home / ".cache" / "paddle" / "dataset").mkdir(parents=True, exist_ok=True)
    os.environ["USERPROFILE"] = str(_paddle_home)
    os.environ["HOME"] = str(_paddle_home)
else:
    # Even on normal systems, create the cache dir to avoid race conditions
    (_user_home / ".cache" / "paddle" / "dataset").mkdir(parents=True, exist_ok=True)
TURKISH_LANG = "tr"
PADDLE_OCR_VERSION = "PP-OCRv5"
MAX_PREPROCESS_PIXELS = 32_000_000
MAX_PADDLE_PAGE_SIDE = 1600
MAX_GRID_ROWS = 80
MAX_GRID_COLS = 36
MAX_GRID_CELLS = 900
REQUIRED_OCR_MODEL_DIRS = ("PP-OCRv5_server_det", "latin_PP-OCRv5_mobile_rec")
REQUIRED_STRUCTURE_MODEL_DIRS = (
    "PP-DocBlockLayout",
    "PP-DocLayout_plus-L",
    "PP-LCNet_x1_0_table_cls",
    "RT-DETR-L_wired_table_cell_det",
    "RT-DETR-L_wireless_table_cell_det",
    "SLANet_plus",
    "SLANeXt_wired",
)


if getattr(sys, "frozen", False):
    if hasattr(sys, "_MEIPASS"):
        TESSERACT_PATH = os.path.join(sys._MEIPASS, "Tesseract-OCR")
    else:
        base_dir = os.path.dirname(sys.executable)
        TESSERACT_PATH = os.path.join(base_dir, "_internal", "Tesseract-OCR")
else:
    TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR"

if not os.path.exists(TESSERACT_PATH):
    TESSERACT_PATH = r"C:\Program Files\Tesseract-OCR"

if pytesseract is not None:
    pytesseract.pytesseract.tesseract_cmd = os.path.join(TESSERACT_PATH, "tesseract.exe")
    if TESSERACT_PATH not in os.environ.get("PATH", ""):
        os.environ["PATH"] += os.pathsep + TESSERACT_PATH

    # --- SUPPRESS TESSERACT CONSOLE WINDOW ---
    if sys.platform == "win32":
        import subprocess
        _orig_popen = pytesseract.pytesseract.subprocess.Popen
        class _PopenNoWindow(_orig_popen):
            def __init__(self, *args, **kwargs):
                if "creationflags" not in kwargs:
                    kwargs["creationflags"] = getattr(subprocess, "CREATE_NO_WINDOW", 0x08000000)
                if "stderr" not in kwargs:
                    kwargs["stderr"] = subprocess.DEVNULL
                super().__init__(*args, **kwargs)
        pytesseract.pytesseract.subprocess.Popen = _PopenNoWindow


def set_tessdata(d):
    """Keep legacy tessdata wiring available for emergency fallback only."""
    if d and os.path.exists(d):
        if os.path.exists(os.path.join(d, "tur.traineddata")):
            os.environ["TESSDATA_PREFIX"] = d
            return True
    return False


tessdata_path = os.path.join(TESSERACT_PATH, "tessdata")
set_tessdata(str(APP_TESSDATA_DIR)) or set_tessdata(tessdata_path)


def get_tesseract_config(psm: int) -> str:
    tessdata_dir = os.environ.get("TESSDATA_PREFIX")
    if not tessdata_dir or not os.path.exists(os.path.join(tessdata_dir, "tur.traineddata")):
        set_tessdata(str(APP_TESSDATA_DIR)) or set_tessdata(tessdata_path)
        tessdata_dir = os.environ.get("TESSDATA_PREFIX", "")

    if tessdata_dir:
        return f'--tessdata-dir {tessdata_dir} --psm {psm} -l tur'
    return f"--psm {psm} -l tur"


@dataclass(slots=True)
class OCRWord:
    text: str
    confidence: float
    box: tuple[int, int, int, int]
    polygon: tuple[tuple[int, int], ...] = field(default_factory=tuple)

    @property
    def cx(self) -> float:
        return (self.box[0] + self.box[2]) / 2

    @property
    def cy(self) -> float:
        return (self.box[1] + self.box[3]) / 2

    @property
    def width(self) -> int:
        return max(1, self.box[2] - self.box[0])

    @property
    def height(self) -> int:
        return max(1, self.box[3] - self.box[1])


@dataclass(slots=True)
class CellSpan:
    row: int
    col: int
    rowspan: int = 1
    colspan: int = 1


@dataclass(slots=True)
class GridDetection:
    h_lines: list[int]
    v_lines: list[int]
    confidence: float


@dataclass(slots=True)
class TableExtractionResult:
    grid: dict[tuple[int, int], str]
    num_rows: int
    num_cols: int
    pipeline_name: str
    table_confidence: float
    ocr_confidence: float
    structure_confidence: float
    spans: list[CellSpan] = field(default_factory=list)
    cell_confidences: dict[tuple[int, int], float] = field(default_factory=dict)

    @property
    def confidence(self) -> float:
        return float(
            0.25 * self.table_confidence
            + 0.40 * self.ocr_confidence
            + 0.35 * self.structure_confidence
        )


def _as_float(value: Any, default: float = 0.0) -> float:
    try:
        if isinstance(value, np.ndarray):
            value = value.item() if value.size == 1 else float(np.mean(value))
        return float(value)
    except Exception:
        return default


def _clip01(value: float) -> float:
    return max(0.0, min(1.0, float(value)))


def configure_paddle_cache(model_storage_dir: str | None = None) -> Path:
    """
    Keep Paddle/PaddleX model files inside the app cache so offline installs and
    packaged builds do not spill model state into arbitrary user directories.
    """
    if model_storage_dir:
        root = Path(model_storage_dir)
        if root.name.lower() == "tessdata":
            root = root.parent / "paddleocr"
    else:
        root = APP_CACHE_DIR / "paddleocr"

    root.mkdir(parents=True, exist_ok=True)
    default_home_cache = Path.home() / ".cache"
    if default_home_cache.exists() and not default_home_cache.is_dir():
        local_home = root / "home"
        (local_home / ".cache" / "paddle").mkdir(parents=True, exist_ok=True)
        os.environ["USERPROFILE"] = str(local_home)
        os.environ["HOME"] = str(local_home)
    os.environ.setdefault("PADDLE_PDX_CACHE_HOME", str(root / "paddlex"))
    os.environ["PADDLE_PDX_DISABLE_MODEL_SOURCE_CHECK"] = "True"
    os.environ["PADDLE_PDX_MODEL_SOURCE"] = os.environ.get("OFFLINEDOCTOOL_PADDLE_MODEL_SOURCE", "bos")
    os.environ["HF_HUB_DISABLE_XET"] = "1"
    os.environ.setdefault("PADDLE_PDX_CPU_NUM_THREADS", str(min(os.cpu_count() or 4, 8)))
    os.environ.setdefault("PADDLEOCR_DISABLE_AUTO_LOGGING_CONFIG", "1")
    return root


def allow_model_downloads() -> bool:
    return os.environ.get("OFFLINEDOCTOOL_ALLOW_MODEL_DOWNLOAD", "").strip().lower() in {"1", "true", "yes", "on"}


def cached_official_model_dir(cache_dir: Path, model_name: str) -> Path:
    return cache_dir / "paddlex" / "official_models" / model_name


def has_cached_official_model(cache_dir: Path, model_name: str) -> bool:
    model_dir = cached_official_model_dir(cache_dir, model_name)
    return (
        model_dir.exists()
        and (model_dir / "inference.yml").exists()
        and ((model_dir / "inference.pdiparams").exists() or (model_dir / "inference.json").exists())
    )


def has_cached_ocr_models(cache_dir: Path) -> bool:
    return all(has_cached_official_model(cache_dir, name) for name in REQUIRED_OCR_MODEL_DIRS)


def has_cached_structure_models(cache_dir: Path) -> bool:
    return all(has_cached_official_model(cache_dir, name) for name in REQUIRED_STRUCTURE_MODEL_DIRS)


class PaddleTurkishOCREngine:
    """Lazy PaddleOCR/PP-Structure wrapper with Turkish-first defaults."""

    def __init__(self, model_storage_dir: str | None = None):
        self.cache_dir = configure_paddle_cache(model_storage_dir)
        self._ocr = None
        self._structure = None
        self._paddle_ocr_cls = None
        self._structure_cls = None
        self.available = False
        self.structure_available = False
        self.error = ""
        self._import_paddle()

    def _import_paddle(self):
        try:
            try:
                import paddlex.utils.deps
                paddlex.utils.deps.require_deps = lambda *args, **kwargs: None
                paddlex.utils.deps.require_extra = lambda *args, **kwargs: None
                paddlex.utils.deps.require_hpip = lambda *args, **kwargs: None
                
                def _safe_is_dep_available(dep, check_version=False):
                    dep_lower = dep.lower()
                    if dep_lower in ("opencv-contrib-python", "opencv-python", "paddlepaddle", "paddlepaddle-gpu", "pyclipper"):
                        return True
                    dep_map = {
                        "opencv-contrib-python": "cv2",
                        "opencv-python": "cv2",
                        "paddlepaddle": "paddle",
                        "paddlepaddle-gpu": "paddle",
                        "pyyaml": "yaml",
                        "pillow": "PIL",
                    }
                    import_name = dep_map.get(dep_lower, dep_lower).replace("-", "_")
                    try:
                        __import__(import_name)
                        return True
                    except ImportError:
                        return False
                        
                paddlex.utils.deps.is_dep_available = _safe_is_dep_available
            except Exception as patch_e:
                print(f"[OCR] Warning: Failed to patch paddlex deps: {patch_e}")

            from paddleocr import PaddleOCR, PPStructureV3

            self._paddle_ocr_cls = PaddleOCR
            self._structure_cls = PPStructureV3
            self.available = True
        except Exception as exc:
            self.error = str(exc)
            print(f"[OCR] PaddleOCR unavailable: {exc}")

    def _common_kwargs(self) -> dict[str, Any]:
        return {
            "lang": TURKISH_LANG,
            "ocr_version": PADDLE_OCR_VERSION,
            "use_doc_orientation_classify": False,
            "use_doc_unwarping": False,
            "use_textline_orientation": False,
            "text_recognition_batch_size": 16,
            "text_det_limit_side_len": 1536,
            "text_det_limit_type": "max",
            "text_det_thresh": 0.20,
            "text_det_box_thresh": 0.35,
            "text_det_unclip_ratio": 1.8,
            "text_rec_score_thresh": 0.03,
            "device": "cpu",
            "enable_mkldnn": False,
            "cpu_threads": min(os.cpu_count() or 4, 8),
        }

    def _load_ocr(self):
        if not self.available:
            return None
        if self._ocr is not None:
            return self._ocr
        if not allow_model_downloads() and not has_cached_ocr_models(self.cache_dir):
            self.error = (
                "PaddleOCR model cache is missing. Set OFFLINEDOCTOOL_ALLOW_MODEL_DOWNLOAD=1 "
                "once or pre-cache PP-OCRv5_server_det and latin_PP-OCRv5_mobile_rec."
            )
            print(f"[OCR] {self.error} Falling back without blocking the UI.")
            return None

        attempts = [
            self._common_kwargs(),
            {
                "lang": TURKISH_LANG,
                "ocr_version": PADDLE_OCR_VERSION,
                "use_doc_orientation_classify": False,
                "use_doc_unwarping": False,
                "use_textline_orientation": False,
                "device": "cpu",
                "enable_mkldnn": False,
            },
            {"lang": TURKISH_LANG, "use_angle_cls": False, "use_gpu": False, "show_log": False},
            {"lang": TURKISH_LANG},
        ]
        last_error = None
        for kwargs in attempts:
            try:
                self._ocr = self._paddle_ocr_cls(**kwargs)
                print("[OCR] PaddleOCR loaded with Turkish-first settings.")
                return self._ocr
            except Exception as exc:
                import traceback
                print(f"[OCR] PaddleOCR init exception:\n{traceback.format_exc()}")
                last_error = exc
        self.available = False
        self.error = str(last_error)
        print(f"[OCR] PaddleOCR load failed: {last_error}")
        return None

    def _load_structure(self):
        if not self.available or self._structure_cls is None:
            return None
        if self._structure is not None:
            return self._structure
        if not allow_model_downloads() and not has_cached_structure_models(self.cache_dir):
            print(
                "[OCR] PP-StructureV3 skipped: table models are not pre-cached. "
                "Set OFFLINEDOCTOOL_ALLOW_MODEL_DOWNLOAD=1 once to cache them."
            )
            return None

        attempts = [
            {
                **self._common_kwargs(),
                "use_table_recognition": True,
                "use_formula_recognition": False,
                "use_chart_recognition": False,
                "use_region_detection": True,
                "format_block_content": True,
            },
            {
                "lang": TURKISH_LANG,
                "ocr_version": PADDLE_OCR_VERSION,
                "use_table_recognition": True,
                "use_doc_orientation_classify": False,
                "use_doc_unwarping": False,
                "use_textline_orientation": False,
                "device": "cpu",
                "enable_mkldnn": False,
            },
        ]
        last_error = None
        for kwargs in attempts:
            try:
                self._structure = self._structure_cls(**kwargs)
                self.structure_available = True
                print("[OCR] PP-StructureV3 loaded.")
                return self._structure
            except Exception as exc:
                last_error = exc
        self.structure_available = False
        print(f"[OCR] PP-StructureV3 unavailable: {last_error}")
        return None

    def recognize_images(self, images: Sequence[np.ndarray]) -> list[tuple[list[OCRWord], float]]:
        if not images:
            return []
        ocr = self._load_ocr()
        if ocr is None:
            return []

        def predict_one(image: np.ndarray) -> tuple[list[OCRWord], float]:
            try:
                raw = ocr.predict(
                    image,
                    use_textline_orientation=False,
                    text_rec_score_thresh=0.03,
                    return_word_box=False,
                )
            except TypeError:
                raw = ocr.ocr(image)
            return parse_paddle_ocr_output(raw)

        try:
            raw_results = ocr.predict(
                list(images),
                use_textline_orientation=False,
                text_rec_score_thresh=0.03,
                return_word_box=False,
            )
            if isinstance(raw_results, list) and len(raw_results) == len(images):
                return [parse_paddle_ocr_output(item) for item in raw_results]
        except Exception as exc:
            print(f"[OCR] Paddle batch OCR fell back to single-image mode: {exc}")

        return [predict_one(image) for image in images]

    def recognize_page(self, image: np.ndarray) -> tuple[list[OCRWord], float]:
        result = self.recognize_images([prepare_page_for_paddle(image)])
        return result[0] if result else ([], 0.0)

    def structure_tables(self, image: np.ndarray) -> list[str]:
        structure = self._load_structure()
        if structure is None:
            return []
        try:
            raw = structure.predict(
                image,
                use_table_recognition=True,
                use_textline_orientation=False,
                use_ocr_results_with_table_cells=True,
                use_wired_table_cells_trans_to_html=True,
                use_wireless_table_cells_trans_to_html=True,
                text_rec_score_thresh=0.03,
            )
            return extract_table_htmls(raw)
        except Exception as exc:
            print(f"[OCR] PP-StructureV3 failed: {exc}")
            return []


_PADDLE_ENGINES: dict[str, PaddleTurkishOCREngine] = {}


def get_paddle_engine(model_storage_dir: str | None = None) -> PaddleTurkishOCREngine:
    key = str(configure_paddle_cache(model_storage_dir))
    engine = _PADDLE_ENGINES.get(key)
    if engine is None:
        engine = PaddleTurkishOCREngine(model_storage_dir)
        _PADDLE_ENGINES[key] = engine
    return engine


class TableTransformerEngine:
    """Lazy Table Transformer engine for robust row/column structure recognition."""

    def __init__(self, model_storage_dir: str | None = None):
        self.model_storage_dir = model_storage_dir
        self.cache_dir = self._configure_cache()
        self.processor = None
        self.model = None
        self.available = False
        self.error = ""
        self._load_model()

    def _configure_cache(self) -> Path:
        if self.model_storage_dir:
            root = Path(self.model_storage_dir)
            if root.name.lower() == "tessdata":
                root = root.parent / "huggingface"
            else:
                root = root / "huggingface"
        else:
            root = APP_CACHE_DIR / "huggingface"
        root.mkdir(parents=True, exist_ok=True)
        os.environ["HF_HOME"] = str(root)
        return root

    def _load_model(self):
        try:
            import torch
            from transformers import AutoImageProcessor, TableTransformerForObjectDetection
            print("[OCR] Loading Table Transformer model (local files preferred)...")
            local_only = not allow_model_downloads()
            
            try:
                self.processor = AutoImageProcessor.from_pretrained(
                    "microsoft/table-transformer-structure-recognition",
                    cache_dir=str(self.cache_dir),
                    local_files_only=local_only
                )
                self.model = TableTransformerForObjectDetection.from_pretrained(
                    "microsoft/table-transformer-structure-recognition",
                    cache_dir=str(self.cache_dir),
                    local_files_only=local_only
                )
            except Exception as e:
                if local_only:
                    print(f"[OCR] Local load failed ({e}). Attempting download since first load...")
                    self.processor = AutoImageProcessor.from_pretrained(
                        "microsoft/table-transformer-structure-recognition",
                        cache_dir=str(self.cache_dir),
                        local_files_only=False
                    )
                    self.model = TableTransformerForObjectDetection.from_pretrained(
                        "microsoft/table-transformer-structure-recognition",
                        cache_dir=str(self.cache_dir),
                        local_files_only=False
                    )
                else:
                    raise e
            self.available = True
            print("[OCR] Table Transformer model loaded successfully.")
        except Exception as exc:
            self.available = False
            self.error = str(exc)
            print(f"[OCR] Table Transformer failed to load: {exc}")


_TATR_ENGINES: dict[str, TableTransformerEngine] = {}


def get_tatr_engine(model_storage_dir: str | None = None) -> TableTransformerEngine:
    if model_storage_dir:
        root = Path(model_storage_dir)
        if root.name.lower() == "tessdata":
            root = root.parent / "huggingface"
        else:
            root = root / "huggingface"
    else:
        root = APP_CACHE_DIR / "huggingface"
    key = str(root.resolve())
    engine = _TATR_ENGINES.get(key)
    if engine is None:
        engine = TableTransformerEngine(model_storage_dir)
        _TATR_ENGINES[key] = engine
    return engine



def _payload_from_result(result: Any) -> Any:
    if isinstance(result, Mapping):
        if "res" in result and isinstance(result["res"], Mapping):
            return result["res"]
        return result
    for attr in ("json", "str"):
        value = getattr(result, attr, None)
        try:
            value = value() if callable(value) else value
        except Exception:
            value = None
        if isinstance(value, Mapping):
            if "res" in value and isinstance(value["res"], Mapping):
                return value["res"]
            return value
    return result


def _first_present(mapping: Mapping[str, Any], names: Sequence[str], default: Any):
    for name in names:
        if name in mapping and mapping[name] is not None:
            return mapping[name]
    return default


def _poly_to_box(poly: Any) -> tuple[int, int, int, int]:
    try:
        arr = np.asarray(poly, dtype=np.float32).reshape(-1, 2)
        x1, y1 = np.min(arr[:, 0]), np.min(arr[:, 1])
        x2, y2 = np.max(arr[:, 0]), np.max(arr[:, 1])
        return int(x1), int(y1), int(x2), int(y2)
    except Exception:
        return 0, 0, 0, 0


def _poly_to_points(poly: Any) -> tuple[tuple[int, int], ...]:
    try:
        arr = np.asarray(poly, dtype=np.float32).reshape(-1, 2)
        return tuple((int(x), int(y)) for x, y in arr)
    except Exception:
        return tuple()


def parse_paddle_ocr_output(result: Any) -> tuple[list[OCRWord], float]:
    """Parse PaddleOCR 3.x dict-like results and older PaddleOCR nested lists."""
    if isinstance(result, list) and result and not isinstance(result[0], (str, bytes)):
        if len(result) == 1:
            return parse_paddle_ocr_output(result[0])
        legacy_words: list[OCRWord] = []
        for item in result:
            if (
                isinstance(item, (list, tuple))
                and len(item) >= 2
                and isinstance(item[1], (list, tuple))
                and len(item[1]) >= 2
            ):
                text = normalize_ocr_text(str(item[1][0]))
                score = _as_float(item[1][1])
                if text:
                    legacy_words.append(OCRWord(text, score, _poly_to_box(item[0]), _poly_to_points(item[0])))
        if legacy_words:
            return legacy_words, float(np.mean([w.confidence for w in legacy_words]))

    payload = _payload_from_result(result)
    if not isinstance(payload, Mapping):
        return [], 0.0

    texts = _first_present(payload, ("rec_texts", "rec_text", "texts"), [])
    scores = _first_present(payload, ("rec_scores", "rec_score", "scores"), [])
    polys = _first_present(payload, ("rec_polys", "dt_polys", "boxes"), [])

    if isinstance(texts, str):
        texts = [texts]
    if isinstance(scores, (float, int, np.floating, np.integer)):
        scores = [scores]

    words: list[OCRWord] = []
    for idx, raw_text in enumerate(list(texts)):
        text = normalize_ocr_text(str(raw_text))
        if not text:
            continue
        score = _as_float(scores[idx], 0.50) if idx < len(scores) else 0.50
        poly = polys[idx] if idx < len(polys) else []
        words.append(OCRWord(text=text, confidence=score, box=_poly_to_box(poly), polygon=_poly_to_points(poly)))

    confidence = float(np.mean([word.confidence for word in words])) if words else 0.0
    return words, confidence


def normalize_ocr_text(text: str) -> str:
    """Unicode-safe cleanup that preserves Turkish characters."""
    if not text:
        return ""

    mojibake = {
        "Ä±": "ı",
        "Ä°": "İ",
        "ÅŸ": "ş",
        "Åž": "Ş",
        "ÄŸ": "ğ",
        "Äž": "Ğ",
        "Ã¼": "ü",
        "Ãœ": "Ü",
        "Ã¶": "ö",
        "Ã–": "Ö",
        "Ã§": "ç",
        "Ã‡": "Ç",
        "â‚º": "₺",
    }
    for bad, good in mojibake.items():
        text = text.replace(bad, good)

    text = unicodedata.normalize("NFC", text)
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t\f\v]+", " ", text)
    text = re.sub(r" *\n *", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = text.strip(" |[]{}()_")
    return text.strip()


def turkish_casefold(text: str) -> str:
    text = unicodedata.normalize("NFC", text)
    text = text.replace("I", "ı").replace("İ", "i")
    return text.lower()


def normalize_header_label(text: str) -> str:
    text = turkish_casefold(normalize_ocr_text(text))
    replacements = {
        "ı": "i",
        "ş": "s",
        "ğ": "g",
        "ü": "u",
        "ö": "o",
        "ç": "c",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return re.sub(r"[^a-z0-9 ]", " ", text).strip()


def looks_like_amount(value: str) -> bool:
    value = normalize_ocr_text(value)
    return bool(
        re.fullmatch(
            r"-?\s*(?:₺|TL)?\s*\d{1,3}(?:[ .]\d{3})*(?:,\d{1,4})?\s*(?:₺|TL)?",
            value,
            flags=re.IGNORECASE,
        )
        or re.fullmatch(r"-?\d+(?:,\d{1,4})?\s*(?:₺|TL)?", value, flags=re.IGNORECASE)
    )


def looks_like_date(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,4}", normalize_ocr_text(value)))


def cleanup_numeric_text(text: str) -> str:
    raw = normalize_ocr_text(text)
    if not raw:
        return raw

    numericish = sum(ch.isdigit() or ch in ".,:%+- ₺TLtl£$€¢" for ch in raw)
    if numericish / max(1, len(raw)) < 0.55:
        return raw

    fixed = raw.translate(str.maketrans({"O": "0", "o": "0", "D": "0"}))
    fixed = re.sub(r"(^|[\s(])(?:t|T|£|\$|€|¢)\s*(?=\d)", r"\1₺", fixed)
    fixed = re.sub(r"(?<=\d)[Il|ı](?=\d|[.,])", "1", fixed)
    fixed = re.sub(r"(?<=\d)S(?=\d|[.,])", "5", fixed)
    fixed = re.sub(r"(?<=\d)B(?=\d|[.,])", "8", fixed)
    fixed = re.sub(r"\s+([.,])", r"\1", fixed)
    fixed = re.sub(r"([.,])\s+", r"\1", fixed)
    fixed = re.sub(r"\bTl\b", "TL", fixed, flags=re.IGNORECASE)
    return fixed


def normalize_date_text(text: str) -> str:
    raw = normalize_ocr_text(text)
    match = re.fullmatch(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", raw)
    if not match:
        return raw
    day, month, year = (int(match.group(1)), int(match.group(2)), match.group(3))
    if len(year) == 2:
        year = "20" + year if int(year) < 50 else "19" + year
    if not (1 <= day <= 31 and 1 <= month <= 12):
        return raw
    return f"{day:02d}.{month:02d}.{year}"


def postprocess_cell_text(text: str, row: int | None = None, col: int | None = None) -> str:
    text = normalize_ocr_text(text)
    if not text:
        return ""

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) > 1:
        deduped = []
        for line in lines:
            if line not in deduped:
                deduped.append(line)
        text = "\n".join(deduped)

    if "\n" not in text:
        numeric = cleanup_numeric_text(text)
        if looks_like_date(numeric):
            return normalize_date_text(numeric)
        if looks_like_amount(numeric):
            return numeric
    return text


def adaptive_upscale(img: np.ndarray, target_min_side: int = 1400) -> np.ndarray:
    h, w = img.shape[:2]
    min_side = min(h, w)
    if min_side <= 0 or min_side >= target_min_side:
        return img
    scale = min(target_min_side / float(min_side), 3.0)
    if h * w * scale * scale > MAX_PREPROCESS_PIXELS:
        scale = math.sqrt(MAX_PREPROCESS_PIXELS / float(h * w))
    if scale <= 1.05:
        return img
    out = cv2.resize(img, (int(round(w * scale)), int(round(h * scale))), interpolation=cv2.INTER_LANCZOS4)
    print(f"[OCR] Adaptive upscale {scale:.2f}x -> {out.shape[1]}x{out.shape[0]}")
    return out


def order_quad_points(points: np.ndarray) -> np.ndarray:
    pts = points.reshape(4, 2).astype("float32")
    rect = np.zeros((4, 2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]
    rect[2] = pts[np.argmax(s)]
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    return rect


def perspective_correct(img: np.ndarray) -> np.ndarray:
    h, w = img.shape[:2]
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blurred, 50, 160)
    edges = cv2.dilate(edges, cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3)), iterations=1)
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    image_area = h * w
    for contour in sorted(contours, key=cv2.contourArea, reverse=True)[:8]:
        area = cv2.contourArea(contour)
        if area < image_area * 0.25:
            continue
        peri = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.025 * peri, True)
        if len(approx) != 4:
            continue

        rect = order_quad_points(approx)
        width_a = np.linalg.norm(rect[2] - rect[3])
        width_b = np.linalg.norm(rect[1] - rect[0])
        height_a = np.linalg.norm(rect[1] - rect[2])
        height_b = np.linalg.norm(rect[0] - rect[3])
        max_w = int(max(width_a, width_b))
        max_h = int(max(height_a, height_b))
        if max_w < w * 0.35 or max_h < h * 0.35:
            continue

        dst = np.array([[0, 0], [max_w - 1, 0], [max_w - 1, max_h - 1], [0, max_h - 1]], dtype="float32")
        matrix = cv2.getPerspectiveTransform(rect, dst)
        warped = cv2.warpPerspective(img, matrix, (max_w, max_h), borderValue=(255, 255, 255))
        print("[OCR] Perspective correction applied.")
        return warped
    return img


def estimate_skew_angle(gray: np.ndarray) -> float:
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)
    min_len = max(40, min(gray.shape[:2]) // 4)
    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold=80, minLineLength=min_len, maxLineGap=12)
    if lines is None:
        return 0.0
    angles = []
    for line in lines[:, 0]:
        x1, y1, x2, y2 = line
        angle = math.degrees(math.atan2(y2 - y1, x2 - x1))
        while angle <= -45:
            angle += 90
        while angle > 45:
            angle -= 90
        if abs(angle) <= 15:
            angles.append(angle)
    if len(angles) < 3:
        return 0.0
    return float(np.median(angles))


def deskew_image(img: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    angle = estimate_skew_angle(gray)
    if abs(angle) < 0.25 or abs(angle) > 12:
        return img
    h, w = img.shape[:2]
    matrix = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
    rotated = cv2.warpAffine(img, matrix, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_CONSTANT, borderValue=(255, 255, 255))
    print(f"[OCR] Deskew applied: {angle:.2f} degrees")
    return rotated


def remove_shadows(img: np.ndarray) -> np.ndarray:
    channels = cv2.split(img)
    normalized = []
    h, w = img.shape[:2]
    kernel = max(21, (min(h, w) // 18) | 1)
    for channel in channels:
        dilated = cv2.dilate(channel, np.ones((7, 7), np.uint8))
        background = cv2.medianBlur(dilated, kernel)
        diff = 255 - cv2.absdiff(channel, background)
        normalized.append(cv2.normalize(diff, None, 0, 255, cv2.NORM_MINMAX))
    return cv2.merge(normalized)


def clahe_sharpen(img: np.ndarray, strong: bool = False) -> np.ndarray:
    lab = cv2.cvtColor(img, cv2.COLOR_BGR2LAB)
    l_channel, a_channel, b_channel = cv2.split(lab)
    clip = 2.6 if strong else 2.0
    clahe = cv2.createCLAHE(clipLimit=clip, tileGridSize=(8, 8))
    l_channel = clahe.apply(l_channel)
    enhanced = cv2.cvtColor(cv2.merge([l_channel, a_channel, b_channel]), cv2.COLOR_LAB2BGR)
    blur_sigma = 1.2 if strong else 0.85
    amount = 1.55 if strong else 1.35
    blurred = cv2.GaussianBlur(enhanced, (0, 0), blur_sigma)
    return cv2.addWeighted(enhanced, amount, blurred, 1 - amount, 0)


def high_contrast_variant(img: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    normalized = cv2.equalizeHist(gray)
    binary = cv2.adaptiveThreshold(
        normalized,
        255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        31,
        9,
    )
    return cv2.cvtColor(binary, cv2.COLOR_GRAY2BGR)


def enhance_image(img: np.ndarray) -> np.ndarray:
    """
    Backward-compatible enhancement entry point, now using the full v11
    document preprocessing stack.
    """
    print(f"[OCR] Original size: {img.shape[1]}x{img.shape[0]}")
    img = perspective_correct(img)
    img = adaptive_upscale(img)
    img = deskew_image(img)
    img = remove_shadows(img)
    img = clahe_sharpen(img)
    print("[OCR] Enhancement done")
    return img


def build_preprocessing_variants(img: np.ndarray) -> list[tuple[str, np.ndarray]]:
    variants: list[tuple[str, np.ndarray]] = [("original", img)]
    corrected = perspective_correct(img)
    balanced = adaptive_upscale(corrected)
    balanced = deskew_image(balanced)
    balanced = remove_shadows(balanced)
    balanced = clahe_sharpen(balanced)
    variants.append(("balanced", balanced))

    if min(img.shape[:2]) < 900 or np.std(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)) < 42:
        variants.append(("high_contrast", high_contrast_variant(balanced)))

    deduped: list[tuple[str, np.ndarray]] = []
    seen: set[tuple[int, int, int]] = set()
    for name, variant in variants:
        key = (variant.shape[0], variant.shape[1], int(np.mean(variant) // 4))
        if key not in seen:
            deduped.append((name, variant))
            seen.add(key)
    return deduped


def merge_nearby_positions(values: Sequence[float | int], tolerance: int = 8) -> list[int]:
    """Merge duplicate line detections that sit on the same visual separator."""
    if not values:
        return []
    merged: list[int] = []
    for value in sorted(int(round(float(item))) for item in values):
        if not merged or value - merged[-1] > tolerance:
            merged.append(value)
        else:
            merged[-1] = int(round((merged[-1] + value) / 2))
    return merged


def _merge_line_components(
    components: Sequence[dict[str, float]],
    tolerance: int = 9,
) -> list[tuple[int, int]]:
    if not components:
        return []

    groups: list[list[dict[str, float]]] = []
    for component in sorted(components, key=lambda item: item["pos"]):
        if not groups:
            groups.append([component])
            continue
        center = float(np.mean([item["pos"] for item in groups[-1]]))
        if abs(component["pos"] - center) <= tolerance:
            groups[-1].append(component)
        else:
            groups.append([component])

    merged: list[tuple[int, int]] = []
    for group in groups:
        weights = [max(1.0, item["length"]) for item in group]
        pos = int(round(np.average([item["pos"] for item in group], weights=weights)))
        intervals = sorted((int(item["start"]), int(item["end"])) for item in group)
        covered: list[list[int]] = []
        for start, end in intervals:
            if not covered or start > covered[-1][1] + tolerance:
                covered.append([start, end])
            else:
                covered[-1][1] = max(covered[-1][1], end)
        coverage = sum(max(0, end - start) for start, end in covered)
        merged.append((pos, coverage))
    return merged


def _component_line_positions(
    mask: np.ndarray,
    orientation: str,
    width: int,
    height: int,
) -> list[int]:
    num_labels, _, stats, _ = cv2.connectedComponentsWithStats(mask, 8)
    components: list[dict[str, float]] = []
    for idx in range(1, num_labels):
        x, y, box_w, box_h, area = [int(v) for v in stats[idx]]
        if orientation == "h":
            if box_w < max(60, width * 0.22):
                continue
            if box_h > max(14, height * 0.045):
                continue
            if area < max(box_w * 0.22, 20):
                continue
            components.append({"pos": y + box_h / 2, "start": x, "end": x + box_w, "length": box_w})
        else:
            if box_h < max(45, height * 0.10):
                continue
            if box_w > max(14, width * 0.035):
                continue
            if area < max(box_h * 0.22, 20):
                continue
            components.append({"pos": x + box_w / 2, "start": y, "end": y + box_h, "length": box_h})

    min_coverage = max(80, width * 0.28) if orientation == "h" else max(70, height * 0.16)
    return [pos for pos, coverage in _merge_line_components(components) if coverage >= min_coverage]


def _projection_line_positions(mask: np.ndarray, orientation: str, width: int, height: int) -> list[int]:
    """
    Projection fallback for line masks where separators are connected to borders.

    Connected-component filtering is excellent at rejecting text strokes, but on
    low-resolution screenshots the full grid can become one connected component.
    A projection over a morphologically-opened line mask still exposes separator
    coordinates without using any document-specific vocabulary.
    """
    projection = np.count_nonzero(mask, axis=1 if orientation == "h" else 0).astype(np.float32)
    if projection.size == 0:
        return []

    smooth = np.convolve(projection, np.ones(3, dtype=np.float32) / 3.0, mode="same")
    threshold = max(width * 0.34, 18) if orientation == "h" else max(height * 0.18, 18)
    positions = find_line_positions(smooth, threshold)

    if orientation == "h":
        return [pos for pos in positions if projection[max(0, pos - 1):pos + 2].max(initial=0) >= max(width * 0.18, 12)]
    return [pos for pos in positions if projection[max(0, pos - 1):pos + 2].max(initial=0) >= max(height * 0.18, 12)]


def _detect_long_line_positions(gray: np.ndarray) -> tuple[list[int], list[int]]:
    """Detect only long table separators; text strokes are intentionally ignored."""
    height, width = gray.shape[:2]
    denoised = cv2.bilateralFilter(gray, 5, 35, 35)
    binary_mean = cv2.adaptiveThreshold(denoised, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 15, 10)
    binary_gauss = cv2.adaptiveThreshold(denoised, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 25, 9)
    binary = cv2.bitwise_or(binary_mean, binary_gauss)
    edges = cv2.Canny(denoised, 25, 110)

    h_component_positions: list[int] = []
    h_projection_positions: list[int] = []
    v_positions: list[int] = []
    h_kernel_sizes = sorted({max(width // 8, 32), max(width // 18, 24)}, reverse=True)
    v_kernel_sizes = sorted({max(height // 8, 24), max(height // 18, 13), max(height // 34, 9)}, reverse=True)

    for source in (binary, edges):
        for kernel_size in h_kernel_sizes:
            horiz = cv2.morphologyEx(
                source,
                cv2.MORPH_OPEN,
                cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_size, 1)),
                iterations=1,
            )
            h_component_positions.extend(_component_line_positions(horiz, "h", width, height))
            h_projection_positions.extend(_projection_line_positions(horiz, "h", width, height))

        for kernel_size in v_kernel_sizes:
            vert = cv2.morphologyEx(
                source,
                cv2.MORPH_OPEN,
                cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_size)),
                iterations=1,
            )
            v_positions.extend(_component_line_positions(vert, "v", width, height))
            v_positions.extend(_projection_line_positions(vert, "v", width, height))

    h_component_positions = merge_nearby_positions(h_component_positions)
    h_projection_positions = merge_nearby_positions(h_projection_positions)
    if len(h_component_positions) >= 3:
        h_positions = list(h_component_positions)
        gaps = np.diff(h_positions)
        useful_gaps = [gap for gap in gaps if gap > 8]
        median_gap = float(np.median(useful_gaps)) if useful_gaps else 0.0
        if median_gap > 0:
            for pos in h_projection_positions:
                for left, right in zip(h_positions, h_positions[1:]):
                    gap = right - left
                    if gap <= median_gap * 1.55:
                        continue
                    if left + median_gap * 0.45 <= pos <= right - median_gap * 0.45:
                        h_positions.append(pos)
                        print(f"[OCR] Added projected horizontal separator at Y={pos}")
                        break
        h_positions = merge_nearby_positions(h_positions)
    else:
        h_positions = h_projection_positions

    # --- Filter out false-positive vertical lines caused by text strokes ---
    # Real grid lines have very uniform pixel intensity along their length
    # (std < 50), while text characters that accidentally align vertically
    # produce high variance (std > 80).  This is universal and does not
    # rely on any table-specific vocabulary.
    merged_v = merge_nearby_positions(v_positions)
    margin = max(int(height * 0.05), 5)
    filtered_v: list[int] = []
    for vx in merged_v:
        if 0 <= vx < width:
            col_strip = gray[margin:height - margin, vx].astype(np.float32)
            col_std = float(np.std(col_strip))
            if col_std <= 50:
                filtered_v.append(vx)
            else:
                print(f"[OCR] Filtered false V-line at X={vx} (std={col_std:.1f}, likely text)")
        else:
            filtered_v.append(vx)

    return merge_nearby_positions(h_positions), filtered_v


def _light_low_saturation_ratio(region: np.ndarray) -> float:
    if region.size == 0:
        return 0.0
    hsv = cv2.cvtColor(region, cv2.COLOR_BGR2HSV)
    saturation = hsv[:, :, 1]
    value = hsv[:, :, 2]
    return float(np.mean((saturation < 15) & (value > 120) & (value < 252)))


def strip_spreadsheet_headers(img: np.ndarray) -> np.ndarray:
    """
    Remove Excel/LibreOffice UI row and column headers from screenshots.

    OCR should see the sheet content, not the A/B/C letters and row numbers.
    """
    h, w = img.shape[:2]
    if h < 140 or w < 180:
        return img

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h_positions, v_positions = _detect_long_line_positions(gray)
    max_header_y = min(90, int(h * 0.12))
    max_header_x = min(90, int(w * 0.10))
    header_y = next((pos for pos in h_positions if 18 <= pos <= max_header_y), None)
    header_x = next((pos for pos in v_positions if 18 <= pos <= max_header_x), None)

    if header_x is None or header_y is None:
        return img

    top_band = img[:header_y, header_x:]
    left_band = img[header_y:, :header_x]
    if _light_low_saturation_ratio(top_band) < 0.45 or _light_low_saturation_ratio(left_band) < 0.45:
        return img

    print(f"[OCR] Removed spreadsheet UI headers: x={header_x}, y={header_y}")
    return img[header_y:, header_x:]


def prepare_page_for_paddle(img: np.ndarray) -> np.ndarray:
    h, w = img.shape[:2]
    max_side = max(h, w)
    if max_side > MAX_PADDLE_PAGE_SIDE:
        scale = MAX_PADDLE_PAGE_SIDE / float(max_side)
        resized = cv2.resize(
            img,
            (max(1, int(round(w * scale))), max(1, int(round(h * scale)))),
            interpolation=cv2.INTER_AREA,
        )
        print(f"[OCR] Downscaled page OCR image to {resized.shape[1]}x{resized.shape[0]}")
        return resized
    return adaptive_upscale(img, target_min_side=1100)


def prepare_cell_for_paddle(cell_img: np.ndarray) -> np.ndarray:
    h, w = cell_img.shape[:2]
    if h < 3 or w < 3:
        return cell_img
    scale = max(1.0, min(3.0, 72.0 / max(h, 1)))
    if scale > 1.05:
        cell_img = cv2.resize(cell_img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_CUBIC)

    gray = cv2.cvtColor(cell_img, cv2.COLOR_BGR2GRAY) if len(cell_img.shape) == 3 else cell_img
    if float(np.mean(gray)) < 130:
        inverted = cv2.bitwise_not(gray)
        clean = cv2.adaptiveThreshold(inverted, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 25, 7)
        cell_img = cv2.cvtColor(clean, cv2.COLOR_GRAY2BGR)
    else:
        cell_img = clahe_sharpen(cell_img if len(cell_img.shape) == 3 else cv2.cvtColor(cell_img, cv2.COLOR_GRAY2BGR))

    return cv2.copyMakeBorder(cell_img, 16, 16, 16, 16, cv2.BORDER_CONSTANT, value=(255, 255, 255))


def find_line_positions(projection, threshold):
    """Extract line center positions from projection array."""
    positions = []
    in_line = False
    start = 0
    for i in range(len(projection)):
        if projection[i] > threshold:
            if not in_line:
                start = i
                in_line = True
        else:
            if in_line:
                positions.append((start + i) // 2)
                in_line = False
    if in_line:
        positions.append((start + len(projection) - 1) // 2)

    if len(positions) > 1:
        merged = [positions[0]]
        for p in positions[1:]:
            if p - merged[-1] > 8:
                merged.append(p)
            else:
                merged[-1] = (merged[-1] + p) // 2
        positions = merged
    return positions


def add_missing_vertical_lines(gray, v_positions):
    """Recover faint column separators missed by edge detection inside unusually wide columns."""
    if len(v_positions) < 3:
        return v_positions

    h, w = gray.shape[:2]
    gaps = [v_positions[i + 1] - v_positions[i] for i in range(len(v_positions) - 1)]
    median_gap = float(np.median(gaps)) if gaps else 0
    if median_gap <= 0:
        return v_positions

    binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 15, 10)
    kernel_h = max(h // 22, 25)
    vert_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_h))
    vert = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vert_kernel, iterations=1)
    projection = np.sum(vert, axis=0)
    candidates = find_line_positions(projection, h * 255 * 0.025)

    refined = list(v_positions)
    min_gap_to_split = max(median_gap * 1.35, median_gap + 45)
    min_side = max(median_gap * 0.32, 22)

    for left, right in zip(v_positions, v_positions[1:]):
        gap = right - left
        if gap < min_gap_to_split:
            continue
        inside = [c for c in candidates if left + min_side <= c <= right - min_side]
        if not inside:
            continue
        best = max(inside, key=lambda c: np.max(projection[max(0, c - 2):min(w, c + 3)]))
        if all(abs(best - x) > 8 for x in refined):
            refined.append(best)
            print(f"[OCR] Added missing vertical separator at X={best}")

    return sorted(refined)


def _slice_has_meaningful_content(gray_slice: np.ndarray) -> bool:
    if gray_slice.size == 0:
        return False

    edges = cv2.Canny(gray_slice, 30, 110)
    num_labels, _, stats, _ = cv2.connectedComponentsWithStats(edges, 8)
    
    component_count = 0
    component_area = 0
    slice_h, slice_w = gray_slice.shape[:2]
    for idx in range(1, num_labels):
        x, y, box_w, box_h, area = [int(v) for v in stats[idx]]
        if area < 8 or box_w < 2 or box_h < 2:
            continue
        if box_w > slice_w * 0.92 or box_h > slice_h * 0.92:
            continue
        component_count += 1
        component_area += area

    return component_count >= 3 or component_area / max(1, gray_slice.size) > 0.005


def _recover_missing_leading_separators(positions: Sequence[int], limit: int) -> list[int]:
    """Recover a first separator when the header and first row were merged."""
    recovered = list(positions)
    if len(recovered) < 4:
        return recovered

    gaps = np.diff(recovered)
    useful_gaps = [gap for gap in gaps if gap > 8]
    if not useful_gaps:
        return recovered

    median_gap = float(np.median(useful_gaps))
    if median_gap <= 8:
        return recovered

    while recovered and recovered[0] > median_gap * 1.45:
        candidate = int(round(recovered[0] - median_gap))
        if candidate <= 6 or candidate >= limit - 1:
            break
        if any(abs(candidate - value) <= 8 for value in recovered):
            break
        recovered.insert(0, candidate)
        print(f"[OCR] Recovered missing leading horizontal separator at Y={candidate}")

    return recovered


def group_close_boxes(boxes, axis="y", tolerance=10):
    """Group rectangles whose centers are close on one axis."""
    if not boxes:
        return []

    index = 1 if axis == "y" else 0
    size_index = 3 if axis == "y" else 2
    sorted_boxes = sorted(boxes, key=lambda box: box[index] + box[size_index] / 2)
    groups = [[sorted_boxes[0]]]

    for box in sorted_boxes[1:]:
        center = box[index] + box[size_index] / 2
        group_centers = [item[index] + item[size_index] / 2 for item in groups[-1]]
        if abs(center - float(np.mean(group_centers))) <= tolerance:
            groups[-1].append(box)
        else:
            groups.append([box])
    return groups


def split_horizontal_clusters(boxes):
    """Split a header row into adjacent table blocks."""
    if not boxes:
        return []

    sorted_boxes = sorted(boxes, key=lambda box: box[0])
    widths = [box[2] for box in sorted_boxes]
    max_gap = min(max(int(np.median(widths) * 0.5), 35), 80)
    clusters = [[sorted_boxes[0]]]

    for box in sorted_boxes[1:]:
        previous = clusters[-1][-1]
        gap = box[0] - (previous[0] + previous[2])
        if gap <= max_gap:
            clusters[-1].append(box)
        else:
            clusters.append([box])
    return clusters


def find_table_bottom(gray, x1, x2, header_y1, header_y2):
    """Find the bottom line after the last row that contains text."""
    h, _ = gray.shape[:2]
    table_slice = gray[:, x1:x2]
    dark = table_slice < 145
    dark_counts = np.sum(dark, axis=1)
    dark_threshold = max(4, int((x2 - x1) * 0.010))
    candidate_rows = np.where(dark_counts > dark_threshold)[0]
    candidate_rows = candidate_rows[candidate_rows > header_y2]

    if len(candidate_rows) == 0:
        return h

    last_text_y = int(candidate_rows[-1])
    binary = cv2.adaptiveThreshold(table_slice, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 15, 10)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max((x2 - x1) // 3, 30), 1))
    horizontal = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel, iterations=1)
    projection = np.sum(horizontal, axis=1)
    lines = find_line_positions(projection, (x2 - x1) * 255 * 0.06)

    for line_y in lines:
        if line_y > last_text_y + 2:
            return min(h, line_y + 2)

    row_height = max(header_y2 - header_y1, 18)
    return min(h, last_text_y + row_height)


def crop_to_main_table(img):
    """Crop Excel-like screenshots down to the main filled table before grid OCR."""
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    hue = hsv[:, :, 0]
    saturation = hsv[:, :, 1]
    value = hsv[:, :, 2]
    green_mask = ((hue >= 35) & (hue <= 95) & (saturation >= 40) & (value >= 60)).astype(np.uint8) * 255
    green_mask = cv2.morphologyEx(
        green_mask,
        cv2.MORPH_CLOSE,
        cv2.getStructuringElement(cv2.MORPH_RECT, (5, 3)),
        iterations=1,
    )
    contours, _ = cv2.findContours(green_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    h, w = img.shape[:2]
    boxes = []
    for contour in contours:
        x, y, box_w, box_h = cv2.boundingRect(contour)
        area = box_w * box_h
        if box_w >= 18 and box_h >= 8 and area >= 140:
            boxes.append((x, y, box_w, box_h))

    if not boxes:
        return img

    row_groups = group_close_boxes(boxes, axis="y", tolerance=max(8, h // 80))
    row_groups = [
        group
        for group in row_groups
        if sum(box[2] for box in group) >= w * 0.18 and max(box[1] for box in group) > h * 0.08
    ]
    if not row_groups:
        return img

    header_group = max(row_groups, key=lambda group: (len(group), sum(box[2] for box in group)))
    clusters = split_horizontal_clusters(header_group)
    if not clusters:
        return img

    main_cluster = max(clusters, key=lambda group: (len(group), sum(box[2] for box in group)))
    x1 = max(0, min(box[0] for box in main_cluster) - 2)
    x2 = min(w, max(box[0] + box[2] for box in main_cluster) + 2)
    header_y1 = max(0, min(box[1] for box in main_cluster) - 2)
    header_y2 = min(h, max(box[1] + box[3] for box in main_cluster) + 2)

    if x2 - x1 < w * 0.2 or header_y2 - header_y1 < 8:
        return img

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    bottom = find_table_bottom(gray, x1, x2, header_y1, header_y2)
    if bottom - header_y1 < (header_y2 - header_y1) * 2:
        return img

    cropped = img[header_y1:bottom, x1:x2]
    print(f"[OCR] Auto-cropped table region: x={x1}:{x2}, y={header_y1}:{bottom}")
    return cropped


def detect_table_grid(img):
    """
    Detect table grid with long-line component filtering.

    The old projection-only detector treated repeated text strokes as grid
    separators. This version keeps only long horizontal/vertical components,
    then adds missing outer boundaries when real content sits outside the
    detected separator range.
    Returns horizontal and vertical separator coordinates.
    """
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    h, w = img.shape[:2]
    h_positions, v_positions = _detect_long_line_positions(gray)
    h_positions = _recover_missing_leading_separators(h_positions, h)

    print(f"[OCR] H-lines before boundary check: {len(h_positions)} at Y={h_positions}")
    print(f"[OCR] V-lines before boundary check: {len(v_positions)} at X={v_positions}")

    if h_positions and h_positions[0] > 12 and _slice_has_meaningful_content(gray[: h_positions[0], :]):
        h_positions.insert(0, 0)
        print("[OCR] Added missing top boundary")

    if h_positions and h - h_positions[-1] > 12 and _slice_has_meaningful_content(gray[h_positions[-1] : h, :]):
        h_positions.append(h - 1)
        print("[OCR] Added missing bottom boundary")

    if v_positions and v_positions[0] > 12 and _slice_has_meaningful_content(gray[:, : v_positions[0]]):
        v_positions.insert(0, 0)
        print("[OCR] Added missing left boundary")

    if v_positions and w - v_positions[-1] > 12 and _slice_has_meaningful_content(gray[:, v_positions[-1] : w]):
        v_positions.append(w - 1)
        print("[OCR] Added missing right boundary")

    h_positions = merge_nearby_positions(h_positions)
    v_positions = merge_nearby_positions(v_positions)
    print(f"[OCR] Final H-lines: {len(h_positions)} at Y={h_positions}")
    print(f"[OCR] Final V-lines: {len(v_positions)} at X={v_positions}")
    return h_positions, v_positions


def evaluate_grid_confidence(img: np.ndarray, h_lines: Sequence[int], v_lines: Sequence[int]) -> float:
    if len(h_lines) < 3 or len(v_lines) < 3:
        return 0.0
    h, w = img.shape[:2]
    row_gaps = np.diff(h_lines)
    col_gaps = np.diff(v_lines)
    if len(row_gaps) == 0 or len(col_gaps) == 0:
        return 0.0
    num_rows = len(h_lines) - 1
    num_cols = len(v_lines) - 1
    num_cells = num_rows * num_cols
    median_row_gap = float(np.median(row_gaps))
    median_col_gap = float(np.median(col_gaps))
    min_reasonable_col = max(24.0, w * 0.018)
    min_reasonable_row = max(14.0, h * 0.006)
    if num_cols > MAX_GRID_COLS or num_rows > MAX_GRID_ROWS or num_cells > MAX_GRID_CELLS:
        print(f"[OCR] Dense grid rejected for per-cell OCR: {num_rows}x{num_cols}")
        return 0.0
    if median_col_gap < min_reasonable_col or median_row_gap < min_reasonable_row:
        print(
            f"[OCR] Dense grid rejected by cell size: "
            f"median row={median_row_gap:.1f}, median col={median_col_gap:.1f}"
        )
        return 0.0

    def regularity(gaps: np.ndarray) -> float:
        median = float(np.median(gaps))
        if median <= 1:
            return 0.0
        return _clip01(1.0 - float(np.std(gaps) / median) * 0.65)

    row_reg = regularity(row_gaps)
    col_reg = regularity(col_gaps)
    coverage = _clip01(((h_lines[-1] - h_lines[0]) / max(1, h) + (v_lines[-1] - v_lines[0]) / max(1, w)) / 2)
    complexity = _clip01(num_cells / 30.0)
    density_penalty = _clip01(max(0, num_cols - 18) / 18.0) * 0.20
    boundary = 0.0
    if h_lines[0] <= max(10, h * 0.02):
        boundary += 0.25
    if h - h_lines[-1] <= max(10, h * 0.02):
        boundary += 0.25
    if v_lines[0] <= max(10, w * 0.02):
        boundary += 0.25
    if w - v_lines[-1] <= max(10, w * 0.02):
        boundary += 0.25
    return _clip01(0.30 * row_reg + 0.30 * col_reg + 0.20 * coverage + 0.12 * complexity + 0.08 * boundary - density_penalty)


def is_reasonable_grid_for_cell_ocr(detection: GridDetection) -> bool:
    rows = len(detection.h_lines) - 1
    cols = len(detection.v_lines) - 1
    cells = rows * cols
    return (
        rows >= 2
        and cols >= 2
        and rows <= MAX_GRID_ROWS
        and cols <= MAX_GRID_COLS
        and cells <= MAX_GRID_CELLS
    )


def detect_table_grid_with_confidence(img: np.ndarray) -> GridDetection:
    h_lines, v_lines = detect_table_grid(img)
    confidence = evaluate_grid_confidence(img, h_lines, v_lines)
    print(f"[OCR] Grid confidence: {confidence:.3f}")
    return GridDetection(h_lines=h_lines, v_lines=v_lines, confidence=confidence)


def words_to_cell_text(words: Sequence[OCRWord]) -> str:
    if not words:
        return ""
    median_h = float(np.median([word.height for word in words])) if words else 12.0
    rows = cluster_items(words, key=lambda word: word.cy, tolerance=max(8.0, median_h * 0.65))
    lines = []
    for row in rows:
        sorted_row = sorted(row, key=lambda word: word.box[0])
        line = " ".join(word.text for word in sorted_row if word.text)
        if line:
            lines.append(line)
    return "\n".join(lines)


def ocr_cells_with_paddle(
    engine: PaddleTurkishOCREngine,
    cell_images: Sequence[np.ndarray],
) -> list[tuple[str, float]]:
    prepared = [prepare_cell_for_paddle(img) for img in cell_images]
    if engine.available:
        results = engine.recognize_images(prepared)
        if len(results) == len(prepared):
            return [(postprocess_cell_text(words_to_cell_text(words)), conf) for words, conf in results]

    return [(ocr_cell_legacy(img), 0.35) for img in cell_images]


def _has_vertical_segment(line_mask: np.ndarray, x: int, y1: int, y2: int) -> bool:
    h, w = line_mask.shape[:2]
    x1, x2 = max(0, x - 2), min(w, x + 3)
    y1, y2 = max(0, y1), min(h, y2)
    if y2 <= y1 or x2 <= x1:
        return True
    roi = line_mask[y1:y2, x1:x2]
    return float(np.count_nonzero(roi)) / max(1, roi.size) > 0.12


def _has_horizontal_segment(line_mask: np.ndarray, y: int, x1: int, x2: int) -> bool:
    h, w = line_mask.shape[:2]
    y1, y2 = max(0, y - 2), min(h, y + 3)
    x1, x2 = max(0, x1), min(w, x2)
    if y2 <= y1 or x2 <= x1:
        return True
    roi = line_mask[y1:y2, x1:x2]
    return float(np.count_nonzero(roi)) / max(1, roi.size) > 0.12


def merge_texts(texts: Iterable[str]) -> str:
    seen = []
    for text in texts:
        text = normalize_ocr_text(text)
        if text and text not in seen:
            seen.append(text)
    return "\n".join(seen)


def detect_cell_spans(
    img: np.ndarray,
    h_lines: Sequence[int],
    v_lines: Sequence[int],
    grid: dict[tuple[int, int], str],
    num_rows: int,
    num_cols: int,
) -> list[CellSpan]:
    """Conservative merged-cell recovery based on missing segment borders."""
    if num_rows <= 0 or num_cols <= 0:
        return []
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY_INV, 15, 10)
    edges = cv2.Canny(gray, 30, 110)
    line_mask = cv2.bitwise_or(binary, edges)
    covered: set[tuple[int, int]] = set()
    spans: list[CellSpan] = []

    for r in range(num_rows):
        c = 0
        while c < num_cols:
            if (r, c) in covered:
                c += 1
                continue
            colspan = 1
            while c + colspan < num_cols:
                boundary_x = v_lines[c + colspan]
                absent = not _has_vertical_segment(line_mask, boundary_x, h_lines[r] + 3, h_lines[r + 1] - 3)
                left_text = grid.get((r, c + colspan - 1), "")
                right_text = grid.get((r, c + colspan), "")
                if not absent:
                    break
                if not left_text and not right_text:
                    break
                colspan += 1

            rowspan = 1
            while r + rowspan < num_rows:
                boundary_y = h_lines[r + rowspan]
                absent = not _has_horizontal_segment(line_mask, boundary_y, v_lines[c] + 3, v_lines[min(num_cols, c + colspan)] - 3)
                below_texts = [grid.get((r + rowspan, cc), "") for cc in range(c, min(num_cols, c + colspan))]
                if not absent or all(not text for text in below_texts):
                    break
                rowspan += 1

            if colspan > 1 or rowspan > 1:
                texts = [
                    grid.get((rr, cc), "")
                    for rr in range(r, min(num_rows, r + rowspan))
                    for cc in range(c, min(num_cols, c + colspan))
                ]
                grid[(r, c)] = merge_texts(texts)
                for rr in range(r, min(num_rows, r + rowspan)):
                    for cc in range(c, min(num_cols, c + colspan)):
                        if (rr, cc) != (r, c):
                            covered.add((rr, cc))
                            grid[(rr, cc)] = ""
                spans.append(CellSpan(r, c, rowspan, colspan))
            c += max(1, colspan)

    return spans


def run_grid_pipeline(
    img: np.ndarray,
    detection: GridDetection,
    engine: PaddleTurkishOCREngine,
    variant_name: str,
) -> TableExtractionResult | None:
    h_lines, v_lines = detection.h_lines, detection.v_lines
    if len(h_lines) < 3 or len(v_lines) < 3:
        return None

    num_rows = len(h_lines) - 1
    num_cols = len(v_lines) - 1
    print(f"[OCR] Grid pipeline on {variant_name}: {num_rows} rows x {num_cols} cols")

    crops: list[np.ndarray] = []
    keys: list[tuple[int, int]] = []
    for r in range(num_rows):
        for c in range(num_cols):
            row_h = h_lines[r + 1] - h_lines[r]
            col_w = v_lines[c + 1] - v_lines[c]
            pad = max(2, int(min(row_h, col_w) * 0.035))
            y1 = h_lines[r] + pad
            y2 = h_lines[r + 1] - pad
            x1 = v_lines[c] + pad
            x2 = v_lines[c + 1] - pad
            if y2 <= y1 or x2 <= x1:
                continue
            crops.append(img[y1:y2, x1:x2])
            keys.append((r, c))

    ocr_results = ocr_cells_with_paddle(engine, crops)
    grid = {(r, c): "" for r in range(num_rows) for c in range(num_cols)}
    cell_confidences: dict[tuple[int, int], float] = {}
    for key, (text, confidence) in zip(keys, ocr_results):
        grid[key] = postprocess_cell_text(text, key[0], key[1])
        cell_confidences[key] = confidence

    for r in range(min(4, num_rows)):
        print(f"[OCR] Row {r}: {[grid.get((r, c), '') for c in range(num_cols)]}")

    clean_rows = finalize_grid(grid, num_rows, num_cols)
    spans = detect_cell_spans(img, h_lines[: clean_rows + 1], v_lines, grid, clean_rows, num_cols)
    ocr_conf = mean_confidence(cell_confidences.values(), default=0.35)
    structure_conf = score_structure(grid, clean_rows, num_cols, detection.confidence)
    return TableExtractionResult(
        grid=grid,
        num_rows=clean_rows,
        num_cols=num_cols,
        pipeline_name=f"opencv_grid+paddle:{variant_name}",
        table_confidence=detection.confidence,
        ocr_confidence=ocr_conf,
        structure_confidence=structure_conf,
        spans=spans,
        cell_confidences=cell_confidences,
    )


def _scaled_lines(lines: Sequence[int], scale: float, max_value: int) -> list[int]:
    scaled = [int(round(value * scale)) for value in lines]
    return [max(0, min(max_value - 1, value)) for value in merge_nearby_positions(scaled)]


def _line_interval_index(value: float, lines: Sequence[int], tolerance: float = 4.0) -> int | None:
    if len(lines) < 2:
        return None
    # Check strict containment first
    for idx in range(len(lines) - 1):
        if lines[idx] <= value <= lines[idx + 1]:
            return idx
    # Check close boundaries (outside the grid)
    if value < lines[0]:
        if lines[0] - value <= max(40.0, tolerance * 2.0):
            return 0
        return None
    if value > lines[-1]:
        if value - lines[-1] <= max(40.0, tolerance * 2.0):
            return len(lines) - 2
        return None
    # Check with tolerance
    for idx in range(len(lines) - 1):
        if lines[idx] - tolerance <= value <= lines[idx + 1] + tolerance:
            return idx
    return None


def _column_for_word(word: OCRWord, v_lines: Sequence[int]) -> int | None:
    x1, _, x2, _ = word.box
    overlaps = []
    for idx in range(len(v_lines) - 1):
        overlap = max(0, min(x2, v_lines[idx + 1]) - max(x1, v_lines[idx]))
        overlaps.append(overlap)
    best = int(np.argmax(overlaps)) if overlaps else None
    if best is not None and overlaps[best] > 0:
        return best
    # Fallback to center
    cx = word.cx
    if cx < v_lines[0]:
        return 0
    if cx > v_lines[-1]:
        return len(v_lines) - 2
    for idx in range(len(v_lines) - 1):
        if v_lines[idx] <= cx <= v_lines[idx + 1]:
            return idx
    return 0


def recover_missing_columns(v_lines: list[int], words: list[OCRWord], h_lines: list[int]) -> list[int]:
    if len(v_lines) < 2 or not words or len(h_lines) < 3:
        return v_lines

    # Exclude words in the header row (between h_lines[0] and h_lines[1])
    header_boundary = h_lines[1]
    data_words = [w for w in words if w.cy > header_boundary]

    if not data_words:
        return v_lines

    # Group data words by row index based on h_lines
    row_words_map = {}
    for w in data_words:
        for r in range(1, len(h_lines) - 1):
            if h_lines[r] <= w.cy <= h_lines[r + 1]:
                row_words_map.setdefault(r, []).append(w)
                break

    num_data_rows = len(row_words_map)
    if num_data_rows < 2:
        return v_lines

    new_v = list(v_lines)

    for i in range(len(v_lines) - 1):
        left = v_lines[i]
        right = v_lines[i + 1]

        # Collect gap intervals in each data row within this column
        gap_intervals = []
        for r, r_words in row_words_map.items():
            col_row_words = [w for w in r_words if left < w.cx < right]
            if len(col_row_words) < 2:
                continue
            
            # Sort words by cx to find gaps in this row
            col_row_words = sorted(col_row_words, key=lambda w: w.cx)
            for k in range(len(col_row_words) - 1):
                w1 = col_row_words[k]
                w2 = col_row_words[k + 1]
                if w2.box[0] > w1.box[2] + 8:
                    gap_intervals.append((w1.box[2], w2.box[0]))

        if not gap_intervals:
            continue

        best_x = None
        max_gaps = 0
        min_intersects = 9999

        # Step through the column width to find the cleanest vertical corridor
        step = max(1, (right - left) // 100)
        for x in range(left + 15, right - 15, step):
            # Count how many rows have a gap containing x
            gap_count = sum(1 for g_start, g_end in gap_intervals if g_start <= x <= g_end)
            
            # Count how many data words overlap with x (allowing 2px margin)
            intersect_count = sum(1 for w in data_words if left < w.cx < right and w.box[0] - 2 < x < w.box[2] + 2)

            # A valid split must be present in a significant fraction of data rows
            if gap_count >= 3 or gap_count >= max(2, num_data_rows // 3):
                if intersect_count <= 1:
                    if intersect_count < min_intersects:
                        min_intersects = intersect_count
                        max_gaps = gap_count
                        best_x = x
                    elif intersect_count == min_intersects and gap_count > max_gaps:
                        max_gaps = gap_count
                        best_x = x

        if best_x is not None:
            new_v.append(best_x)
            print(f"[OCR] Recovered missing column separator at X={best_x} inside column [{left}, {right}] with {max_gaps} gaps")

    return sorted(list(set(new_v)))


def _split_known_header_overflow(text: str, col: int, num_cols: int) -> list[tuple[int, str]] | None:
    normalized = normalize_header_label(text)
    compact = normalized.replace(" ", "")
    if "hesapbakiyesi" in compact and "genelbakiye" in compact and col + 1 < num_cols:
        return [(col, "Hesap Bakiyesi"), (col + 1, "Genel Bakiye")]
    return None


def split_word_by_spaces(word: OCRWord) -> list[OCRWord]:
    text = word.text
    if not text or " " not in text:
        return [word]
    parts = text.split(" ")
    parts = [p for p in parts if p]
    if len(parts) <= 1:
        return [word]
    tokens = []
    x1, y1, x2, y2 = word.box
    total_len = len(text)
    char_width = (x2 - x1) / max(1, total_len)
    current_char_idx = 0
    for part in parts:
        start_idx = text.find(part, current_char_idx)
        if start_idx == -1:
            start_idx = current_char_idx
        token_len = len(part)
        token_x1 = int(round(x1 + start_idx * char_width))
        token_x2 = int(round(token_x1 + token_len * char_width))
        token_x1 = max(x1, min(x2, token_x1))
        token_x2 = max(x1, min(x2, token_x2))
        token_word = OCRWord(
            text=part,
            confidence=word.confidence,
            box=(token_x1, y1, token_x2, y2)
        )
        tokens.append(token_word)
        current_char_idx = start_idx + token_len
    return tokens


def run_grid_geometry_pipeline(
    img: np.ndarray,
    detection: GridDetection,
    engine: PaddleTurkishOCREngine,
    variant_name: str,
    precomputed_words: list = None,
) -> TableExtractionResult | None:
    """Use page OCR boxes but place them into the detected spreadsheet grid."""
    if not engine.available or len(detection.h_lines) < 3 or len(detection.v_lines) < 3:
        return None

    prepared = prepare_page_for_paddle(img)
    if precomputed_words is not None:
        words = precomputed_words
        page_conf = float(np.mean([w.confidence for w in words])) if words else 0.85
    else:
        recognized = engine.recognize_images([prepared])
        if not recognized:
            return None
        words, page_conf = recognized[0]
        words = [word for word in words if word.text and word.confidence >= 0.10 and not is_garbage_text(word.text)]

    if len(words) < 4:
        return None

    scale_x = prepared.shape[1] / max(1, img.shape[1])
    scale_y = prepared.shape[0] / max(1, img.shape[0])
    h_lines = _scaled_lines(detection.h_lines, scale_y, prepared.shape[0])
    v_lines = _scaled_lines(detection.v_lines, scale_x, prepared.shape[1])
    v_lines = recover_missing_columns(v_lines, words, h_lines)
    if len(h_lines) < 3 or len(v_lines) < 3:
        return None

    num_rows = len(h_lines) - 1
    num_cols = len(v_lines) - 1
    grouped: dict[tuple[int, int], list[OCRWord]] = {}
    direct_text: dict[tuple[int, int], list[str]] = {}
    confidences: dict[tuple[int, int], list[float]] = {}

    for word in words:
        row = _line_interval_index(word.cy, h_lines, tolerance=max(6.0, word.height * 0.35))
        if row is None or row >= num_rows:
            continue
        col = _column_for_word(word, v_lines)
        if col is None or col >= num_cols:
            continue

        split_header = _split_known_header_overflow(word.text, col, num_cols)
        if split_header:
            for split_col, split_text in split_header:
                key = (row, split_col)
                direct_text.setdefault(key, []).append(split_text)
                confidences.setdefault(key, []).append(word.confidence)
            continue

        key = (row, col)
        grouped.setdefault(key, []).append(word)
        confidences.setdefault(key, []).append(word.confidence)

    if len(grouped) + len(direct_text) < 4:
        return None

    grid = {(r, c): "" for r in range(num_rows) for c in range(num_cols)}
    cell_confidences: dict[tuple[int, int], float] = {}
    for key, cell_words in grouped.items():
        grid[key] = postprocess_cell_text(words_to_cell_text(cell_words), key[0], key[1])
    for key, parts in direct_text.items():
        grid[key] = merge_texts([grid.get(key, ""), *parts])
    for key, values in confidences.items():
        cell_confidences[key] = mean_confidence(values, default=page_conf)

    clean_rows = finalize_grid(grid, num_rows, num_cols)
    structure_conf = score_structure(grid, clean_rows, num_cols, detection.confidence)
    return TableExtractionResult(
        grid=grid,
        num_rows=clean_rows,
        num_cols=num_cols,
        pipeline_name=f"opencv_grid_geometry+paddle:{variant_name}",
        table_confidence=max(detection.confidence, 0.70),
        ocr_confidence=page_conf,
        structure_confidence=max(structure_conf, 0.70),
        cell_confidences=cell_confidences,
    )


def cluster_items(items: Sequence[Any], key, tolerance: float) -> list[list[Any]]:
    if not items:
        return []
    ordered = sorted(items, key=key)
    clusters: list[list[Any]] = [[ordered[0]]]
    centers = [float(key(ordered[0]))]
    for item in ordered[1:]:
        value = float(key(item))
        if abs(value - float(np.mean(centers))) <= tolerance:
            clusters[-1].append(item)
            centers.append(value)
        else:
            clusters.append([item])
            centers = [value]
    return clusters


def cluster_positions(values: Sequence[float], tolerance: float) -> list[float]:
    if not values:
        return []
    clusters = cluster_items(list(values), key=lambda value: value, tolerance=tolerance)
    return [float(np.mean(cluster)) for cluster in clusters]


def infer_borderless_table(
    img: np.ndarray,
    engine: PaddleTurkishOCREngine,
    variant_name: str,
) -> TableExtractionResult | None:
    words, page_conf = engine.recognize_page(img)
    words = [word for word in words if word.text and word.confidence >= 0.10 and not is_garbage_text(word.text)]
    if len(words) < 4:
        return None

    median_h = float(np.median([word.height for word in words]))
    row_groups = cluster_items(words, key=lambda word: word.cy, tolerance=max(10.0, median_h * 0.75))
    if len(row_groups) < 2:
        return None

    logical_cells = []
    for row_idx, row_words in enumerate(row_groups):
        sorted_words = sorted(row_words, key=lambda word: word.box[0])
        if not sorted_words:
            continue
        current = [sorted_words[0]]
        for word in sorted_words[1:]:
            previous = current[-1]
            gap = word.box[0] - previous.box[2]
            threshold = max(median_h * 2.2, np.median([w.width for w in row_words]) * 0.60, 26)
            if gap > threshold:
                logical_cells.append((row_idx, current))
                current = [word]
            else:
                current.append(word)
        logical_cells.append((row_idx, current))

    if len(logical_cells) < 4:
        return None

    lefts = [min(word.box[0] for word in group) for _, group in logical_cells]
    col_centers = cluster_positions(lefts, tolerance=max(28.0, median_h * 2.4))
    if len(col_centers) < 2:
        return None

    grid: dict[tuple[int, int], str] = {}
    cell_confidences: dict[tuple[int, int], float] = {}
    for row_idx, group in logical_cells:
        left = min(word.box[0] for word in group)
        col_idx = min(range(len(col_centers)), key=lambda idx: abs(col_centers[idx] - left))
        key = (row_idx, col_idx)
        text = words_to_cell_text(group)
        confidence = float(np.mean([word.confidence for word in group]))
        if key in grid and grid[key]:
            grid[key] = merge_texts([grid[key], text])
            cell_confidences[key] = max(cell_confidences.get(key, 0.0), confidence)
        else:
            grid[key] = postprocess_cell_text(text, row_idx, col_idx)
            cell_confidences[key] = confidence

    num_rows = len(row_groups)
    num_cols = len(col_centers)
    for r in range(num_rows):
        for c in range(num_cols):
            grid.setdefault((r, c), "")

    clean_rows = finalize_grid(grid, num_rows, num_cols)
    filled = sum(1 for value in grid.values() if value.strip())
    table_conf = _clip01(0.35 + min(num_rows * num_cols, 40) / 80.0)
    structure_conf = _clip01(0.40 + min(filled / max(1, clean_rows * num_cols), 0.75))
    return TableExtractionResult(
        grid=grid,
        num_rows=clean_rows,
        num_cols=num_cols,
        pipeline_name=f"borderless_paddle:{variant_name}",
        table_confidence=table_conf,
        ocr_confidence=page_conf,
        structure_confidence=structure_conf,
        cell_confidences=cell_confidences,
    )


class HTMLTableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables: list[list[list[dict[str, Any]]]] = []
        self._current_table: list[list[dict[str, Any]]] | None = None
        self._current_row: list[dict[str, Any]] | None = None
        self._current_cell: dict[str, Any] | None = None
        self._table_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        tag = tag.lower()
        if tag == "table":
            self._table_depth += 1
            if self._table_depth == 1:
                self._current_table = []
        elif tag == "tr" and self._current_table is not None:
            self._current_row = []
        elif tag in ("td", "th") and self._current_row is not None:
            attr_map = {name.lower(): value for name, value in attrs}
            self._current_cell = {
                "text": [],
                "rowspan": max(1, int(attr_map.get("rowspan") or 1)),
                "colspan": max(1, int(attr_map.get("colspan") or 1)),
            }
        elif tag == "br" and self._current_cell is not None:
            self._current_cell["text"].append("\n")

    def handle_data(self, data: str) -> None:
        if self._current_cell is not None:
            self._current_cell["text"].append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in ("td", "th") and self._current_cell is not None and self._current_row is not None:
            self._current_cell["text"] = normalize_ocr_text(" ".join(self._current_cell["text"]))
            self._current_row.append(self._current_cell)
            self._current_cell = None
        elif tag == "tr" and self._current_row is not None and self._current_table is not None:
            self._current_table.append(self._current_row)
            self._current_row = None
        elif tag == "table":
            if self._table_depth == 1 and self._current_table is not None:
                self.tables.append(self._current_table)
                self._current_table = None
            self._table_depth = max(0, self._table_depth - 1)


def table_from_html(html: str, pipeline_name: str = "pp_structure") -> TableExtractionResult | None:
    parser = HTMLTableParser()
    try:
        parser.feed(html)
    except Exception as exc:
        print(f"[OCR] HTML table parse failed: {exc}")
        return None
    if not parser.tables:
        return None

    raw_table = max(parser.tables, key=lambda rows: sum(len(row) for row in rows))
    grid: dict[tuple[int, int], str] = {}
    spans: list[CellSpan] = []
    occupied: set[tuple[int, int]] = set()
    max_col = 0
    for r, row in enumerate(raw_table):
        c = 0
        for cell in row:
            while (r, c) in occupied:
                c += 1
            text = postprocess_cell_text(str(cell.get("text", "")), r, c)
            rowspan = int(cell.get("rowspan", 1))
            colspan = int(cell.get("colspan", 1))
            grid[(r, c)] = text
            if rowspan > 1 or colspan > 1:
                spans.append(CellSpan(r, c, rowspan, colspan))
            for rr in range(r, r + rowspan):
                for cc in range(c, c + colspan):
                    occupied.add((rr, cc))
                    grid.setdefault((rr, cc), "")
            max_col = max(max_col, c + colspan)
            c += colspan

    num_rows = max((r for r, _ in grid.keys()), default=-1) + 1
    num_cols = max_col
    if num_rows < 1 or num_cols < 2:
        return None
    for r in range(num_rows):
        for c in range(num_cols):
            grid.setdefault((r, c), "")

    clean_rows = finalize_grid(grid, num_rows, num_cols)
    filled = sum(1 for value in grid.values() if value.strip())
    structure_conf = _clip01(0.55 + min(filled / max(1, clean_rows * num_cols), 0.40))
    return TableExtractionResult(
        grid=grid,
        num_rows=clean_rows,
        num_cols=num_cols,
        pipeline_name=pipeline_name,
        table_confidence=0.82,
        ocr_confidence=0.72,
        structure_confidence=structure_conf,
        spans=spans,
    )


def extract_table_htmls(result: Any) -> list[str]:
    htmls: list[str] = []

    def walk(value: Any, depth: int = 0) -> None:
        if depth > 8:
            return
        if isinstance(value, str):
            if "<table" in value.lower() and "</table" in value.lower():
                htmls.append(value)
            return
        if isinstance(value, Mapping):
            for item in value.values():
                walk(item, depth + 1)
            return
        if isinstance(value, (list, tuple)):
            for item in value:
                walk(item, depth + 1)
            return
        for attr in ("html", "json", "str", "xlsx"):
            obj = getattr(value, attr, None)
            if obj is None:
                continue
            try:
                obj = obj() if callable(obj) else obj
            except Exception:
                continue
            walk(obj, depth + 1)

    walk(result)
    unique = []
    for html in htmls:
        if html not in unique:
            unique.append(html)
    return unique


def run_table_transformer_pipeline(
    img: np.ndarray,
    tatr_engine: TableTransformerEngine,
    engine: PaddleTurkishOCREngine,
    variant_name: str,
) -> list[TableExtractionResult]:
    if not tatr_engine.available or not engine.available:
        return []

    try:
        from PIL import Image as PILImage
        import torch

        update_progress(22, f"Table Transformer ile tablo sınırları analiz ediliyor ('{variant_name}' varyantı)...")
        pil_img = PILImage.fromarray(cv2.cvtColor(img, cv2.COLOR_BGR2RGB))
        W, H = pil_img.size

        update_progress(25, f"Table Transformer yapay zeka modeli çalıştırılıyor ('{variant_name}' varyantı)...")
        inputs = tatr_engine.processor(images=pil_img, return_tensors="pt")
        with torch.no_grad():
            outputs = tatr_engine.model(**inputs)

        target_sizes = torch.tensor([[H, W]])
        results = tatr_engine.processor.post_process_object_detection(
            outputs, threshold=0.3, target_sizes=target_sizes
        )[0]

        id2label = tatr_engine.model.config.id2label

        cols = []
        rows = []

        for score, label, box in zip(results["scores"], results["labels"], results["boxes"]):
            label_name = id2label[label.item()]
            score_val = score.item()
            box_coords = box.tolist()

            if label_name == "table column":
                cols.append((box_coords, score_val))
            elif label_name in ["table row", "table column header"]:
                rows.append((box_coords, score_val))

        def merge_intervals(boxes, horizontal=True):
            if not boxes:
                return []
            idx = 0 if horizontal else 1
            end_idx = 2 if horizontal else 3
            sorted_boxes = sorted(boxes, key=lambda b: b[0][idx])
            merged = []
            for box, score in sorted_boxes:
                if not merged:
                    merged.append((box, score))
                else:
                    last_box, last_score = merged[-1]
                    span_current = box[end_idx] - box[idx]
                    span_last = last_box[end_idx] - last_box[idx]
                    overlap = max(0, min(box[end_idx], last_box[end_idx]) - max(box[idx], last_box[idx]))
                    if overlap > 0.6 * min(span_current, span_last):
                        new_box = list(last_box)
                        new_box[idx] = min(last_box[idx], box[idx])
                        new_box[end_idx] = max(last_box[end_idx], box[end_idx])
                        merged[-1] = (new_box, max(last_score, score))
                    else:
                        merged.append((box, score))
            return merged

        merged_cols = merge_intervals(cols, horizontal=True)
        merged_rows = merge_intervals(rows, horizontal=False)

        # Safe fix: merge abnormally narrow TATR columns/rows that are likely false splits
        def _merge_narrow_detections(detections, axis_start, axis_end, min_count=3, ratio=0.45):
            """Merge detections narrower than ratio * median width with their closest neighbor."""
            if len(detections) <= min_count:
                return detections
            changed = True
            while changed:
                changed = False
                if len(detections) <= min_count:
                    break
                sizes = [d[0][axis_end] - d[0][axis_start] for d in detections]
                median_size = float(sorted(sizes)[len(sizes) // 2])
                min_size = median_size * ratio
                for i in range(len(detections)):
                    box, score = detections[i]
                    size = box[axis_end] - box[axis_start]
                    if size < min_size:
                        if i == 0:
                            merge_with = 1
                        elif i == len(detections) - 1:
                            merge_with = i - 1
                        else:
                            gap_before = box[axis_start] - detections[i - 1][0][axis_end]
                            gap_after = detections[i + 1][0][axis_start] - box[axis_end]
                            merge_with = i - 1 if gap_before <= gap_after else i + 1
                        nb_box, nb_score = detections[merge_with]
                        new_box = [
                            min(box[0], nb_box[0]), min(box[1], nb_box[1]),
                            max(box[2], nb_box[2]), max(box[3], nb_box[3]),
                        ]
                        lo, hi = min(i, merge_with), max(i, merge_with)
                        detections[lo] = (new_box, max(score, nb_score))
                        detections.pop(hi)
                        changed = True
                        break
            return detections

        merged_cols = _merge_narrow_detections(merged_cols, axis_start=0, axis_end=2)
        merged_rows = _merge_narrow_detections(merged_rows, axis_start=1, axis_end=3)

        # Second pass: merge adjacent column pairs where BOTH are narrow and
        # essentially touching (tiny gap) — a clear sign TATR split one real column in two
        if len(merged_cols) > 3:
            sizes = [d[0][2] - d[0][0] for d in merged_cols]
            median_size = float(sorted(sizes)[len(sizes) // 2])
            i = 0
            while i < len(merged_cols) - 1 and len(merged_cols) > 3:
                box1, s1 = merged_cols[i]
                box2, s2 = merged_cols[i + 1]
                w1 = box1[2] - box1[0]
                w2 = box2[2] - box2[0]
                gap = max(0, box2[0] - box1[2])
                # Both narrow (<65% of median) AND gap is tiny (<10% of median)
                if w1 < median_size * 0.65 and w2 < median_size * 0.65 and gap < median_size * 0.10:
                    new_box = [min(box1[0], box2[0]), min(box1[1], box2[1]),
                               max(box1[2], box2[2]), max(box1[3], box2[3])]
                    merged_cols[i] = (new_box, max(s1, s2))
                    merged_cols.pop(i + 1)
                    sizes = [d[0][2] - d[0][0] for d in merged_cols]
                    median_size = float(sorted(sizes)[len(sizes) // 2])
                else:
                    i += 1

        if len(merged_cols) < 2 or len(merged_rows) < 2:
            print(f"[OCR] TATR found too few columns/rows (cols: {len(merged_cols)}, rows: {len(merged_rows)})")
            return []

        v_lines = []
        v_lines.append(int(round(merged_cols[0][0][0])))
        for i in range(len(merged_cols) - 1):
            mid = (merged_cols[i][0][2] + merged_cols[i+1][0][0]) / 2
            v_lines.append(int(round(mid)))
        v_lines.append(int(round(merged_cols[-1][0][2])))

        h_lines = []
        h_lines.append(int(round(merged_rows[0][0][1])))
        for i in range(len(merged_rows) - 1):
            mid = (merged_rows[i][0][3] + merged_rows[i+1][0][1]) / 2
            h_lines.append(int(round(mid)))
        h_lines.append(int(round(merged_rows[-1][0][3])))

        v_lines = [max(0, min(W - 1, x)) for x in v_lines]
        h_lines = [max(0, min(H - 1, y)) for y in h_lines]

        v_lines = merge_nearby_positions(v_lines, tolerance=8)
        h_lines = merge_nearby_positions(h_lines, tolerance=6)

        # Robust grid alignment: Clip lines to word boundaries to prevent empty column/row overflows
        update_progress(35, f"Izgara hizalaması için sayfa metinleri analiz ediliyor ('{variant_name}' varyantı)...")
        prepared_for_adjust = prepare_page_for_paddle(img)
        update_progress(40, f"Türkçe OCR motoru ile sayfa metinleri okunuyor ('{variant_name}' varyantı - Bu işlem 10 saniye sürebilir)...")
        recognized_for_adjust = engine.recognize_images([prepared_for_adjust])
        words_filtered = None
        if recognized_for_adjust:
            words_for_adjust, _ = recognized_for_adjust[0]
            words_filtered = []
            scale_x_adj = img.shape[1] / max(1, prepared_for_adjust.shape[1])
            scale_y_adj = img.shape[0] / max(1, prepared_for_adjust.shape[0])
            adj_words = []
            for w in words_for_adjust:
                if w.text and w.confidence >= 0.10 and not is_garbage_text(w.text):
                    words_filtered.append(w)
                    box_img = (
                        int(round(w.box[0] * scale_x_adj)),
                        int(round(w.box[1] * scale_y_adj)),
                        int(round(w.box[2] * scale_x_adj)),
                        int(round(w.box[3] * scale_y_adj)),
                    )
                    adj_words.append(OCRWord(text=w.text, confidence=w.confidence, box=box_img))
            
            if adj_words:
                # 1D clustering on word cy values to find text row centers
                cy_vals = sorted([w.cy for w in adj_words])
                clusters = []
                if cy_vals:
                    current_cluster = [cy_vals[0]]
                    for val in cy_vals[1:]:
                        if val - current_cluster[-1] < 12: # 12px threshold for same row
                            current_cluster.append(val)
                        else:
                            clusters.append(float(np.mean(current_cluster)))
                            current_cluster = [val]
                    clusters.append(float(np.mean(current_cluster)))

                # Ensure boundaries cover the word extent
                all_x1 = [w.box[0] for w in adj_words]
                all_x2 = [w.box[2] for w in adj_words]
                min_x = max(0, min(all_x1))
                max_x = min(W - 1, max(all_x2))

                # Refine h_lines
                refined_h = list(h_lines)
                if not any(abs(y - 0) <= 6 for y in refined_h):
                    refined_h.insert(0, 0)
                if not any(abs(y - (H - 1)) <= 6 for y in refined_h):
                    refined_h.append(H - 1)
                refined_h = sorted(list(set(refined_h)))

                # Remove intermediate lines cutting directly through text row centers
                i_adj = 1
                while i_adj < len(refined_h) - 1:
                    line_val = refined_h[i_adj]
                    too_close = False
                    for c_val in clusters:
                        if abs(c_val - line_val) < 8:
                            too_close = True
                            break
                    if too_close:
                        refined_h.pop(i_adj)
                    else:
                        i_adj += 1

                # Ensure there is at least one separator between consecutive text row clusters
                for idx_c in range(len(clusters) - 1):
                    c1_val = clusters[idx_c]
                    c2_val = clusters[idx_c + 1]
                    has_sep = any(c1_val < y < c2_val for y in refined_h)
                    if not has_sep:
                        mid_val = int(round((c1_val + c2_val) / 2))
                        refined_h.append(mid_val)

                # Ensure top and bottom coverage
                if clusters:
                    if not any(y < clusters[0] for y in refined_h):
                        refined_h.insert(0, 0)
                    if not any(y > clusters[-1] for y in refined_h):
                        refined_h.append(H - 1)

                h_lines = sorted(list(set(refined_h)))
                h_lines = merge_nearby_positions(h_lines, tolerance=6)

                # Refine v_lines: nudge each internal line to nearest text gap
                refined_v = list(v_lines)
                if not any(abs(x - min_x) <= 8 for x in refined_v):
                    refined_v.insert(0, int(min_x))
                if not any(abs(x - max_x) <= 8 for x in refined_v):
                    refined_v.append(int(max_x))
                refined_v = sorted(list(set(refined_v)))

                # Safe nudge: for each internal v_line, find the nearest gap between words
                avg_col_w = (refined_v[-1] - refined_v[0]) / max(1, len(refined_v) - 1)
                nudge_radius = avg_col_w * 0.40  # nudge within 40% of avg column width
                for vi in range(1, len(refined_v) - 1):
                    vx = refined_v[vi]
                    # Collect right-edges of words to the left, left-edges of words to the right
                    left_wall = vx - nudge_radius
                    right_wall = vx + nudge_radius
                    rights_of_left = []  # right edges of words LEFT of vx
                    lefts_of_right = []  # left edges of words RIGHT of vx
                    for w in adj_words:
                        wcx = w.cx
                        if wcx < vx and w.box[2] > left_wall:
                            rights_of_left.append(w.box[2])
                        elif wcx > vx and w.box[0] < right_wall:
                            lefts_of_right.append(w.box[0])
                    if rights_of_left and lefts_of_right:
                        gap_left = max(rights_of_left)
                        gap_right = min(lefts_of_right)
                        if gap_right > gap_left + 2:  # real gap exists
                            refined_v[vi] = int(round((gap_left + gap_right) / 2))

                v_lines = sorted(list(set(refined_v)))
                v_lines = merge_nearby_positions(v_lines, tolerance=8)
                v_lines = recover_missing_columns(v_lines, adj_words, h_lines)

        if len(h_lines) < 3 or len(v_lines) < 3:
            return []

        detection = GridDetection(h_lines=h_lines, v_lines=v_lines, confidence=0.88)
        candidates = []

        update_progress(70, f"Izgara geometrisi modeline göre hücre metinleri eşleştiriliyor ('{variant_name}' varyantı)...")
        geo_candidate = run_grid_geometry_pipeline(img, detection, engine, f"tatr_{variant_name}", precomputed_words=words_filtered)
        if geo_candidate:
            geo_candidate.table_confidence = 0.88
            geo_candidate.structure_confidence = score_structure(
                geo_candidate.grid, geo_candidate.num_rows, geo_candidate.num_cols, 0.88
            )
            candidates.append(geo_candidate)

        total_cells = (len(h_lines) - 1) * (len(v_lines) - 1)
        if total_cells < 80:
            update_progress(80, f"Tablo hücreleri tek tek analiz ediliyor ('{variant_name}' varyantı)...")
            cell_candidate = run_grid_pipeline(img, detection, engine, f"tatr_{variant_name}")
            if cell_candidate:
                cell_candidate.table_confidence = 0.88
                cell_candidate.structure_confidence = score_structure(
                    cell_candidate.grid, cell_candidate.num_rows, cell_candidate.num_cols, 0.88
                )
                candidates.append(cell_candidate)

        return candidates

    except Exception as e:
        print(f"[OCR] Error in TATR pipeline: {e}")
        return []


def run_pp_structure_pipeline(
    img: np.ndarray,

    engine: PaddleTurkishOCREngine,
    variant_name: str,
) -> TableExtractionResult | None:
    htmls = engine.structure_tables(img)
    candidates = [
        table_from_html(html, pipeline_name=f"pp_structure_v3:{variant_name}")
        for html in htmls
    ]
    candidates = [candidate for candidate in candidates if candidate is not None]
    if not candidates:
        return None
    best = max(candidates, key=lambda candidate: candidate.confidence)
    print(f"[OCR] PP-Structure candidate confidence: {best.confidence:.3f}")
    return best


def ocr_cell_legacy(cell_img: np.ndarray) -> str:
    """Emergency fallback only when PaddleOCR is unavailable."""
    if pytesseract is None:
        return ""
    h, w = cell_img.shape[:2]
    if h < 3 or w < 3:
        return ""
    scale = max(1, 120 // max(h, 1))
    if scale > 1:
        cell_img = cv2.resize(cell_img, (w * scale, h * scale), interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(cell_img, cv2.COLOR_BGR2GRAY) if len(cell_img.shape) == 3 else cell_img
    mean_val = np.mean(gray)
    if mean_val < 140:
        inverted = cv2.bitwise_not(gray)
        clean = cv2.adaptiveThreshold(inverted, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 21, 5)
        if np.mean(clean) < 128:
            clean = cv2.bitwise_not(clean)
    else:
        _, clean = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        if np.mean(clean) < 128:
            clean = cv2.bitwise_not(clean)
    padded = cv2.copyMakeBorder(clean, 20, 20, 20, 20, cv2.BORDER_CONSTANT, value=255)
    pil = PILImage.fromarray(padded)
    try:
        text = pytesseract.image_to_string(pil, config=f"{get_tesseract_config(7)} --oem 1").strip()
        return postprocess_cell_text(text)
    except Exception as exc:
        print(f"[OCR] Legacy Tesseract cell fallback error: {exc}")
        return ""


def ocr_cell(cell_img: np.ndarray) -> str:
    """Backward-compatible public cell OCR helper."""
    engine = get_paddle_engine()
    result = ocr_cells_with_paddle(engine, [cell_img])
    return result[0][0] if result else ""


def is_garbage_text(text):
    """Check if text looks like OCR garbage, without penalizing Turkish letters."""
    if not text:
        return False
    alnum = sum(1 for c in text if c.isalnum())
    noise = sum(1 for c in text if c in "—–«»‘’“”;:!?@#$%^&*~`<>=")
    total = len(text.strip())
    if total == 0:
        return False
    if total > 3 and noise / total > 0.45 and alnum < 2:
        return True
    if total <= 2 and noise > 0 and alnum == 0:
        return True
    return False


def filter_garbage_rows(grid, num_rows, num_cols):
    """
    Remove trailing rows that are mostly empty or OCR noise.
    The header row is never removed.
    """
    clean_rows = num_rows
    for r in range(num_rows - 1, 0, -1):
        cells = [grid.get((r, c), "") for c in range(num_cols)]
        filled = sum(1 for c in cells if c.strip())
        garbage = sum(1 for c in cells if is_garbage_text(c))
        meaningful = any(
            looks_like_amount(c)
            or looks_like_date(c)
            or sum(1 for ch in c if ch.isalnum()) >= 3
            for c in cells
        )
        if (
            filled == 0
            or garbage >= max(2, math.ceil(num_cols * 0.6))
            or (filled <= 1 and garbage > 0 and not meaningful)
        ):
            print(f"[OCR] Removing garbage row {r}: {cells}")
            for c in range(num_cols):
                grid.pop((r, c), None)
            clean_rows = r
        else:
            break
    return clean_rows


HEADER_CANONICALS = [
    "Satış Temsilcisi",
    "Müşteri Adı",
    "Müşteri Soyadı",
    "Müşteri Ad",
    "Müşteri Soyad",
    "Satış Bölgesi",
    "Ürün Adı",
    "Ürün Ad",
    "Satış Tarihi",
    "Tarih",
    "Satış Tutarı",
    "Tutar",
    "Tarih Kategori",
    "Kategori",
    "Gelir",
    "Gider",
    "Hesap Bakiyesi",
    "Genel Bakiye",
    "Fatura No",
    "Fatura Tarihi",
    "Firma",
    "Firma Ünvanı",
    "Vergi No",
    "Vergi Dairesi",
    "Adres",
    "Telefon",
    "E-posta",
    "Açıklama",
    "Birim Fiyat",
    "Miktar",
    "KDV",
    "İskonto",
    "Ara Toplam",
    "Genel Toplam",
    "Toplam",
]

HEADER_ALIASES = {
    "satis temsilcisi": "Satış Temsilcisi",
    "temsilci": "Satış Temsilcisi",
    "musteri adi": "Müşteri Adı",
    "musteri ad": "Müşteri Adı",
    "musteri soyadi": "Müşteri Soyadı",
    "musteri soyad": "Müşteri Soyadı",
    "satis bolgesi": "Satış Bölgesi",
    "bolge": "Satış Bölgesi",
    "urun adi": "Ürün Adı",
    "urun ad": "Ürün Adı",
    "urun": "Ürün Adı",
    "satis tarihi": "Satış Tarihi",
    "tarih": "Tarih",
    "fatura tarihi": "Fatura Tarihi",
    "satis tutari": "Satış Tutarı",
    "tutar": "Tutar",
    "kategori": "Kategori",
    "gelir": "Gelir",
    "gider": "Gider",
    "hesap bakiyesi": "Hesap Bakiyesi",
    "genel bakiye": "Genel Bakiye",
    "toplam": "Toplam",
    "genel toplam": "Genel Toplam",
    "kdv": "KDV",
    "birim fiyat": "Birim Fiyat",
    "miktar": "Miktar",
    "fatura no": "Fatura No",
    "fis no": "Fatura No",
    "vergi no": "Vergi No",
    "vkn": "Vergi No",
    "tckn": "Vergi No",
    "vergi dairesi": "Vergi Dairesi",
    "firma unvani": "Firma Ünvanı",
    "firma": "Firma",
    "adres": "Adres",
    "telefon": "Telefon",
    "eposta": "E-posta",
    "e posta": "E-posta",
    "aciklama": "Açıklama",
}


def normalize_generic_table_layout(result: TableExtractionResult) -> bool:
    """Small layout cleanup that does not depend on a known table template.

    Previously merged sparse first rows into a 'title' cell.  This was removed
    because it destroyed real header rows in non-standard tables.  The function
    is kept as a no-op hook for future safe cleanups.
    """
    return False


def enhance_selected_result(result: TableExtractionResult) -> TableExtractionResult:
    normalize_generic_table_layout(result)
    return result


def infer_header_from_values(values):
    values = [normalize_ocr_text(value) for value in values if normalize_ocr_text(value)]
    if not values:
        return ""
    amount_count = sum(1 for value in values if looks_like_amount(value))
    date_count = sum(1 for value in values if looks_like_date(value))
    email_count = sum(1 for value in values if re.search(r"@.+\.", value))
    phone_count = sum(1 for value in values if re.fullmatch(r"[+() 0-9-]{7,}", value))
    tax_count = sum(1 for value in values if re.fullmatch(r"\d{10,11}", re.sub(r"\D", "", value)))
    quantity_count = sum(1 for value in values if re.fullmatch(r"\d+(?:,\d+)?", value))
    threshold = max(2, math.ceil(len(values) * 0.45))

    if date_count >= threshold:
        return "Tarih"
    if amount_count >= threshold:
        return "Tutar"
    if email_count >= threshold:
        return "E-posta"
    if phone_count >= threshold:
        return "Telefon"
    if tax_count >= threshold:
        return "Vergi No"
    if quantity_count >= threshold:
        return "Miktar"
    return ""


def is_probable_title_row(grid: Mapping[tuple[int, int], str], num_rows: int, num_cols: int) -> bool:
    if num_rows < 2 or num_cols < 2:
        return False
    first_filled = [normalize_ocr_text(grid.get((0, col), "")) for col in range(num_cols)]
    second_filled = [normalize_ocr_text(grid.get((1, col), "")) for col in range(num_cols)]
    first_filled = [value for value in first_filled if value]
    second_filled = [value for value in second_filled if value]
    return (
        bool(first_filled)
        and len(first_filled) <= 2
        and len(second_filled) >= max(2, math.ceil(num_cols * 0.50))
    )


def repair_headers(grid, num_rows, num_cols, header_row: int = 0):
    """Universal header repair: only fix Turkish diacritics OCR errors.

    Does NOT replace headers via dictionaries or guess from column values.
    Whatever OCR reads from the image is preserved as-is, except for
    common Turkish character OCR mistakes (s→ş, u→ü, i→ı, o→ö, c→ç, g→ğ).
    """
    if num_rows <= header_row:
        return

    # Common OCR diacritics substitution pairs for Turkish
    _TURKISH_DIACRITIC_FIXES = {
        "s": "ş", "S": "Ş",
        "u": "ü", "U": "Ü",
        "i": "ı", "I": "İ",
        "o": "ö", "O": "Ö",
        "c": "ç", "C": "Ç",
        "g": "ğ", "G": "Ğ",
    }

    canonical_map = {normalize_header_label(h): h for h in HEADER_CANONICALS}

    for col in range(num_cols):
        header = grid.get((header_row, col), "").strip()
        if not header:
            continue

        normalized = normalize_header_label(header)

        # Only attempt a fix if the normalized form is a very close match
        # to a known canonical header (cutoff 0.85 = almost identical,
        # just missing Turkish diacritics).
        # This prevents "Oca 2022" from matching "Açıklama" or similar.
        if normalized in canonical_map:
            fixed = canonical_map[normalized]
            if fixed != header:
                grid[(header_row, col)] = fixed
                print(f"[OCR] Fixed diacritics at column {col}: '{header}' -> '{fixed}'")


def finalize_grid(grid: dict[tuple[int, int], str], num_rows: int, num_cols: int) -> int:
    for r in range(num_rows):
        for c in range(num_cols):
            grid[(r, c)] = postprocess_cell_text(grid.get((r, c), ""), r, c)
    clean_rows = filter_garbage_rows(grid, num_rows, num_cols)
    for r in range(clean_rows):
        for c in range(num_cols):
            grid.setdefault((r, c), "")



    header_row = 0
    repair_headers(grid, clean_rows, num_cols, header_row=header_row)
    return clean_rows


def mean_confidence(values: Iterable[float], default: float = 0.0) -> float:
    vals = [_clip01(float(v)) for v in values if v is not None]
    return float(np.mean(vals)) if vals else default


def score_structure(
    grid: Mapping[tuple[int, int], str],
    num_rows: int,
    num_cols: int,
    base_confidence: float,
) -> float:
    if num_rows <= 0 or num_cols <= 0:
        return 0.0
    total = num_rows * num_cols
    filled = sum(1 for r in range(num_rows) for c in range(num_cols) if grid.get((r, c), "").strip())
    filled_ratio = filled / max(1, total)
    header_fill = sum(1 for c in range(num_cols) if grid.get((0, c), "").strip()) / max(1, num_cols)
    garbage = sum(1 for value in grid.values() if is_garbage_text(value))
    garbage_penalty = min(0.25, garbage / max(1, total))
    return _clip01(0.45 * base_confidence + 0.25 * min(filled_ratio * 1.8, 1.0) + 0.25 * header_fill + 0.05 * min(num_cols / 4, 1.0) - garbage_penalty)


def choose_best_result(candidates: Sequence[TableExtractionResult]) -> TableExtractionResult | None:
    valid = [candidate for candidate in candidates if candidate.num_rows > 0 and candidate.num_cols > 0]
    if not valid:
        return None
    valid.sort(key=lambda candidate: (candidate.confidence, candidate.num_rows * candidate.num_cols), reverse=True)
    best = valid[0]
    grid_candidates = [
        candidate
        for candidate in valid
        if candidate.pipeline_name.startswith("opencv_grid")
        and candidate.table_confidence >= 0.58
        and candidate.num_cols >= 2
    ]
    if grid_candidates:
        cell_grid_candidates = [
            candidate
            for candidate in grid_candidates
            if candidate.pipeline_name.startswith("opencv_grid+paddle")
        ]
        if cell_grid_candidates:
            cell_grid_candidates.sort(
                key=lambda candidate: (
                    candidate.table_confidence,
                    candidate.structure_confidence,
                    candidate.confidence,
                    candidate.num_rows * candidate.num_cols,
                ),
                reverse=True,
            )
            best_cell_grid = cell_grid_candidates[0]
            if best_cell_grid.table_confidence >= 0.62:
                best = best_cell_grid

        grid_candidates.sort(
            key=lambda candidate: (
                candidate.table_confidence,
                candidate.structure_confidence,
                candidate.confidence,
                candidate.num_rows * candidate.num_cols,
            ),
            reverse=True,
        )
        best_grid = grid_candidates[0]
        if (
            not best.pipeline_name.startswith("opencv_grid+paddle")
            and (best_grid.table_confidence >= 0.62 or best_grid.confidence >= best.confidence - 0.12)
        ):
            best = best_grid

    print("[OCR] Candidate scores:")
    for candidate in valid:
        print(
            f"[OCR]   {candidate.pipeline_name}: total={candidate.confidence:.3f}, "
            f"table={candidate.table_confidence:.3f}, ocr={candidate.ocr_confidence:.3f}, "
            f"structure={candidate.structure_confidence:.3f}"
        )
    print(f"[OCR] Selected pipeline: {best.pipeline_name}")
    return best


def write_excel(
    grid,
    num_rows,
    num_cols,
    output_path,
    spans: Sequence[CellSpan] | None = None,
    cell_confidences: Mapping[tuple[int, int], float] | None = None,
    pipeline_name: str | None = None,
    stage_confidences: Mapping[str, float] | None = None,
):
    """Write grid to styled Excel with merged cells, wrapping, and Turkish-safe text."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tablo"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    cell_font = Font(name="Calibri", size=11)
    border = Border(
        left=Side(style="thin", color="A6A6A6"),
        right=Side(style="thin", color="A6A6A6"),
        top=Side(style="thin", color="A6A6A6"),
        bottom=Side(style="thin", color="A6A6A6"),
    )
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_rows = {0}
    if spans and any(span.row == 0 and span.col == 0 and span.colspan >= num_cols for span in spans):
        header_rows.add(1)

    for r in range(num_rows):
        for c in range(num_cols):
            value = normalize_ocr_text(str(grid.get((r, c), "")))
            cell = ws.cell(row=r + 1, column=c + 1, value=value)
            cell.font = header_font if r in header_rows else cell_font
            cell.border = border
            cell.alignment = align
            if r in header_rows:
                cell.fill = header_fill

    if spans:
        for span in spans:
            if span.row >= num_rows or span.col >= num_cols:
                continue
            end_row = min(num_rows, span.row + max(1, span.rowspan))
            end_col = min(num_cols, span.col + max(1, span.colspan))
            if end_row - span.row > 1 or end_col - span.col > 1:
                ws.merge_cells(
                    start_row=span.row + 1,
                    start_column=span.col + 1,
                    end_row=end_row,
                    end_column=end_col,
                )

    for c in range(1, num_cols + 1):
        max_len = 0
        for r in range(1, num_rows + 1):
            value = ws.cell(row=r, column=c).value
            if value:
                max_len = max(max_len, max(len(part) for part in str(value).splitlines()))
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = min(max(max_len + 4, 12), 42)

    for r in range(1, num_rows + 1):
        max_lines = 1
        for c in range(1, num_cols + 1):
            value = ws.cell(row=r, column=c).value
            if value:
                max_lines = max(max_lines, str(value).count("\n") + 1)
        ws.row_dimensions[r].height = max(18, min(90, 17 * max_lines))

    if num_rows >= 1 and num_cols >= 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    ws.sheet_view.showGridLines = True
    if pipeline_name or stage_confidences:
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output_path)
    wb.close()
    print(f"[OCR] Excel saved: {output_path}")


def run_legacy_fallback(img: np.ndarray) -> TableExtractionResult | None:
    """Last-resort fallback kept only to avoid total failure if Paddle models are missing."""
    if pytesseract is None:
        return None
    h, w = img.shape[:2]
    scale = 3 if max(h, w) < 800 else 2
    up = cv2.resize(img, (w * scale, h * scale), interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(up, cv2.COLOR_BGR2GRAY)
    pil = PILImage.fromarray(gray)

    try:
        data = pytesseract.image_to_data(pil, config=get_tesseract_config(6), output_type=pytesseract.Output.DICT)
    except Exception as exc:
        print(f"[OCR] Legacy fallback failed: {exc}")
        return None

    words = []
    for i in range(len(data["text"])):
        raw = normalize_ocr_text(data["text"][i].strip())
        conf = int(float(data["conf"][i]))
        if conf < 20 or not raw or all(c in "|[]{}()_" for c in raw):
            continue
        words.append(
            {
                "text": raw.strip("|[]{}()_"),
                "left": data["left"][i],
                "cy": data["top"][i] + data["height"][i] // 2,
                "height": data["height"][i],
                "conf": conf / 100.0,
            }
        )

    if not words:
        return None

    med_h = np.median([word["height"] for word in words])

    def cluster(vals, tol):
        if not vals:
            return []
        indexed = sorted(enumerate(vals), key=lambda x: x[1])
        clusters = [([indexed[0][0]], [indexed[0][1]])]
        for idx, value in indexed[1:]:
            if abs(value - np.mean(clusters[-1][1])) <= tol:
                clusters[-1][0].append(idx)
                clusters[-1][1].append(value)
            else:
                clusters.append(([idx], [value]))
        return [(int(np.mean(values)), ids) for ids, values in clusters]

    rows_cl = sorted(cluster([word["cy"] for word in words], med_h * 0.40), key=lambda x: x[0])
    cols_cl = sorted(cluster([word["left"] for word in words], med_h * 1.25), key=lambda x: x[0])
    if len(rows_cl) < 1 or len(cols_cl) < 2:
        return None

    word_rows, word_cols = {}, {}
    for row_idx, (_, ids) in enumerate(rows_cl):
        for word_idx in ids:
            word_rows[word_idx] = row_idx
    for col_idx, (_, ids) in enumerate(cols_cl):
        for word_idx in ids:
            word_cols[word_idx] = col_idx

    grouped: dict[tuple[int, int], list[tuple[int, str]]] = {}
    confidences: dict[tuple[int, int], float] = {}
    for word_idx, word in enumerate(words):
        row_idx, col_idx = word_rows.get(word_idx), word_cols.get(word_idx)
        if row_idx is None or col_idx is None:
            continue
        key = (row_idx, col_idx)
        grouped.setdefault(key, []).append((word["left"], word["text"]))
        confidences[key] = max(confidences.get(key, 0.0), word["conf"])

    grid = {}
    for key, word_list in grouped.items():
        word_list.sort(key=lambda item: item[0])
        grid[key] = postprocess_cell_text(" ".join(word for _, word in word_list), key[0], key[1])

    num_rows, num_cols = len(rows_cl), len(cols_cl)
    for r in range(num_rows):
        for c in range(num_cols):
            grid.setdefault((r, c), "")
    clean_rows = finalize_grid(grid, num_rows, num_cols)
    return TableExtractionResult(
        grid=grid,
        num_rows=clean_rows,
        num_cols=num_cols,
        pipeline_name="legacy_tesseract_emergency",
        table_confidence=0.35,
        ocr_confidence=mean_confidence(confidences.values(), default=0.35),
        structure_confidence=score_structure(grid, clean_rows, num_cols, 0.35),
        cell_confidences=confidences,
    )


current_progress = {"percent": 0, "status": "Hazırlanıyor..."}

def update_progress(percent: int, status: str):
    global current_progress
    if percent < current_progress.get("percent", 0) and percent > 0:
        percent = current_progress.get("percent", 0)
    current_progress["percent"] = percent
    current_progress["status"] = status
    print(f"[OCR PROGRESS] {percent}% - {status}")

def extract_table_to_excel(image_path: str, output_path: str, model_storage_dir: str = None):
    """Main entry point."""
    update_progress(5, "OCR motorları yükleniyor...")
    print("[OCR] === v11 START (PaddleOCR Turkish-first) ===")
    if not os.path.exists(image_path):
        raise ValueError(f"Not found: {image_path}")

    set_tessdata(model_storage_dir)
    engine = get_paddle_engine(model_storage_dir)
    tatr_engine = get_tatr_engine(model_storage_dir)

    update_progress(15, "Görüntü ön işleme ve iyileştirme yapılıyor...")
    img = cv2.imread(image_path)
    if img is None:
        raise ValueError(f"Cannot read: {image_path}")

    img = strip_spreadsheet_headers(img)
    h, w = img.shape[:2]
    print(f"[OCR] Image: {w}x{h}")

    candidates: list[TableExtractionResult] = []
    variants = build_preprocessing_variants(img)
    high_confidence_seen = False

    if tatr_engine.available:
        print("[OCR] Table Transformer engine is available. Running TATR pipeline...")
        for idx, (variant_name, variant_img) in enumerate(variants):
            pct = 20 + int((idx / len(variants)) * 30)
            update_progress(pct, f"Table Transformer ile tablo yapısı analiz ediliyor ('{variant_name}' varyantı)...")
            table_img = crop_to_main_table(variant_img)
            tatr_candidates = run_table_transformer_pipeline(table_img, tatr_engine, engine, variant_name)
            if tatr_candidates:
                candidates.extend(tatr_candidates)
                if any(c.confidence >= 0.82 for c in tatr_candidates):
                    high_confidence_seen = True
                    break

    if not high_confidence_seen:
        for idx, (variant_name, variant_img) in enumerate(variants):
            pct = 50 + int((idx / len(variants)) * 30)
            update_progress(pct, f"OpenCV ile ızgara yapısı ve kenarlıklar aranıyor ('{variant_name}' varyantı)...")
            table_img = crop_to_main_table(variant_img)
            th, tw = table_img.shape[:2]
            print(f"[OCR] Variant '{variant_name}' OCR region: {tw}x{th}")

            detection = detect_table_grid_with_confidence(table_img)
            if detection.confidence >= 0.34 and is_reasonable_grid_for_cell_ocr(detection):
                update_progress(pct + 3, f"Hücre metinleri Türkçe OCR ile okunuyor ('{variant_name}' varyantı)...")
                geometry_candidate = run_grid_geometry_pipeline(table_img, detection, engine, variant_name)
                if geometry_candidate:
                    candidates.append(geometry_candidate)
                candidate = run_grid_pipeline(table_img, detection, engine, variant_name)
                if candidate:
                    candidates.append(candidate)
                    high_confidence_seen = candidate.confidence >= 0.82
            elif len(detection.h_lines) >= 3 and len(detection.v_lines) >= 3:
                print(
                    f"[OCR] Grid pipeline skipped: low confidence or too dense "
                    f"({len(detection.h_lines) - 1}x{len(detection.v_lines) - 1})."
                )

            if engine.available and (not high_confidence_seen or detection.confidence < 0.62):
                update_progress(pct + 6, f"Çerçevesiz tablo analizi yapılıyor ('{variant_name}' varyantı)...")
                borderless = infer_borderless_table(table_img, engine, variant_name)
                if borderless:
                    candidates.append(borderless)
                    high_confidence_seen = borderless.confidence >= 0.62

            if high_confidence_seen:
                break

    best = choose_best_result(candidates)

    if engine.available and (best is None or best.confidence < 0.72):
        update_progress(82, "Düşük doğruluk, PP-StructureV3 algoritması ile alternatif analiz yapılıyor...")
        print("[OCR] Low confidence; trying PP-StructureV3 table recognition.")
        for variant_name, variant_img in variants[:2]:
            table_img = crop_to_main_table(variant_img)
            pp_candidate = run_pp_structure_pipeline(table_img, engine, variant_name)
            if pp_candidate:
                candidates.append(pp_candidate)
        best = choose_best_result(candidates)

    if best is None:
        update_progress(87, "Yedek OCR sistemi ile metin çıkarımı yapılıyor...")
        print("[OCR] Paddle pipelines unavailable or empty; using emergency legacy fallback.")
        fallback = run_legacy_fallback(crop_to_main_table(enhance_image(img)))
        if fallback:
            best = fallback

    if best is None:
        raise ValueError("No table text found")

    update_progress(92, "Hücre birleştirmeleri ve doğruluk analizi sonlandırılıyor...")
    best = enhance_selected_result(best)

    print(
        f"[OCR] Final confidence: total={best.confidence:.3f}, "
        f"table={best.table_confidence:.3f}, ocr={best.ocr_confidence:.3f}, "
        f"structure={best.structure_confidence:.3f}"
    )
    print(f"[OCR] Clean table: {best.num_rows} rows x {best.num_cols} cols")

    update_progress(96, "Excel tablosu oluşturuluyor ve biçimlendiriliyor...")
    write_excel(
        best.grid,
        best.num_rows,
        best.num_cols,
        output_path,
        spans=best.spans,
        cell_confidences=best.cell_confidences,
        pipeline_name=best.pipeline_name,
        stage_confidences={
            "table": best.table_confidence,
            "ocr": best.ocr_confidence,
            "structure": best.structure_confidence,
        },
    )
    update_progress(100, "Tamamlandı!")
    print(f"[OCR] === DONE {best.pipeline_name} ({best.num_rows}x{best.num_cols}) ===")
    return best.num_rows, best.num_cols
